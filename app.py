import json
import os
import re
import socket
import threading
import webbrowser
from datetime import datetime, date
from calendar import monthrange
import csv
import io
import sqlite3
from urllib.parse import urlencode
from urllib.parse import parse_qs
from jinja2 import Template
from fastapi import FastAPI, Request, HTTPException
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, Response
import uvicorn
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

DATA_FILE = "debts.json"
DB_FILE = os.environ.get("DEBT_TRACKER_DB", "debt_tracker.db")

app = FastAPI()


def route(path, methods):
    return app.api_route(path, methods=methods, response_class=HTMLResponse)


app.route = route


def abort(status_code):
    raise HTTPException(status_code=status_code)


def redirect(url, status_code=302):
    return RedirectResponse(url, status_code=status_code)


def url_for(endpoint, **values):
    path_params = {}
    query_params = {}
    for key, value in values.items():
        try:
            app.url_path_for(endpoint, **{key: value})
            path_params[key] = value
        except Exception:
            query_params[key] = value
    url = str(app.url_path_for(endpoint, **path_params))
    if query_params:
        url = f"{url}?{urlencode(query_params, doseq=True)}"
    return url


def render_template_string(template, **context):
    context.setdefault('url_for', url_for)
    context.setdefault('request', None)
    return Template(template).render(**context)


async def read_form_data(request: Request):
    try:
        return await request.form()
    except AssertionError:
        body = await request.body()
        parsed = parse_qs(body.decode('utf-8'))
        return {k: v[-1] if v else '' for k, v in parsed.items()}


def jsonify(data):
    return JSONResponse(data)


def get_db_connection():
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    return conn


def init_storage():
    conn = get_db_connection()
    try:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS app_state (
                id INTEGER PRIMARY KEY CHECK (id = 1),
                payload TEXT NOT NULL
            )
            """
        )
        conn.commit()
        row = conn.execute("SELECT payload FROM app_state WHERE id = 1").fetchone()
        if row is None:
            if os.path.exists(DATA_FILE):
                try:
                    with open(DATA_FILE, "r", encoding="utf-8") as f:
                        payload = f.read().strip() or json.dumps({"accounts": []})
                except OSError:
                    payload = json.dumps({"accounts": []})
            else:
                payload = json.dumps({"accounts": []})
            conn.execute("INSERT INTO app_state (id, payload) VALUES (1, ?)", (payload,))
            conn.commit()
    finally:
        conn.close()


def excel_safe(value):
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, (dict, list, tuple, set)):
        return json.dumps(value, ensure_ascii=False, default=str)
    return value


def sheet_from_rows(workbook, title, rows):
    sheet = workbook.create_sheet(title=title)
    rows = list(rows)
    if not rows:
        sheet.append(["No data"])
        return sheet
    headers = list(rows[0].keys())
    sheet.append(headers)
    for row in rows:
        sheet.append([excel_safe(row.get(header, "")) for header in headers])
    return sheet


def style_header_row(sheet, row_number=1, fill="1F4E78", font_color="FFFFFF"):
    for cell in sheet[row_number]:
        cell.font = Font(bold=True, color=font_color)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = Alignment(horizontal="center", vertical="center")


def autofit_sheet(sheet, max_width=40):
    for column_cells in sheet.columns:
        lengths = []
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            lengths.append(len(value))
        if not lengths:
            continue
        width = min(max(max(lengths) + 2, 10), max_width)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width


def coerce_excel_value(value):
    if value is None:
      return None
    if isinstance(value, str):
      text = value.strip()
      if text == '':
        return None
      if text.lower() in ('true', 'false'):
        return text.lower() == 'true'
      return text
    if isinstance(value, (datetime, date)):
      return value.isoformat()
    return value

TEMPLATE = """
<!doctype html>
<html lang="en">
  <head>
    <meta charset="utf-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1" />
    <title>Debt Tracker</title>
    <style>
      :root {
        --bg: #f7f9fc;
        --surface: #ffffff;
        --surface-2: #f8fafc;
        --border: #e2e8f0;
        --text: #1b2330;
        --text-muted: #475569;
        --text-dim: #64748b;
        --summary-bg: #eef2ff;
        --nav-active-bg: #eff6ff;
        --nav-active-color: #2563eb;
        --cancel-bg: #e5e7eb;
        --cancel-text: #1f2937;
        --modal-overlay: rgba(0,0,0,0.4);
        --success-bg: #ecfdf5;
        --success-text: #166534;
        --code-bg: #f1f5f9;
        --chart-grid: #e2e8f0;
        --chart-label: #1f2937;
        --chart-axis: #475569;
        --chart-no-data: #64748b;
        --account-link-color: #111;
      }
      [data-theme="dark"] {
        --bg: #0f172a;
        --surface: #1e293b;
        --surface-2: #131f36;
        --border: #334155;
        --text: #f1f5f9;
        --text-muted: #94a3b8;
        --text-dim: #94a3b8;
        --summary-bg: #1a2744;
        --nav-active-bg: rgba(59,130,246,0.15);
        --nav-active-color: #3b82f6;
        --cancel-bg: #334155;
        --cancel-text: #f1f5f9;
        --modal-overlay: rgba(0,0,0,0.7);
        --success-bg: rgba(16,185,129,0.15);
        --success-text: #34d399;
        --code-bg: #0f172a;
        --chart-grid: #334155;
        --chart-label: #f1f5f9;
        --chart-axis: #94a3b8;
        --chart-no-data: #94a3b8;
        --account-link-color: #f1f5f9;
      }
      body { font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Arial, sans-serif; margin: 2rem; background: var(--bg); color: var(--text); transition: background 0.2s, color 0.2s; }
      h1 { margin-bottom: 0.5rem; }
      .card { background: var(--surface); padding: 1.25rem; border-radius: 0.625rem; box-shadow: 0 8px 24px rgba(15,23,42,0.08); max-width: 780px; margin: 0 auto; border: 1px solid var(--border); }
      [data-theme="dark"] .card { box-shadow: 0 8px 32px rgba(0,0,0,0.5); }
      .account-row { display: grid; grid-template-columns: 2fr 1fr 1fr 0.8fr; gap: 0.5rem; align-items: center; margin-bottom: 0.5rem; padding: 0.45rem 0.55rem; border-radius:6px; background: var(--surface-2); border: 1px solid var(--border); }
      .account-row label { font-weight: 600; }
      .account-row input { padding: 0.55rem 0.65rem; border: 1px solid var(--border); border-radius: 0.5rem; width: 100%; background: var(--bg); color: var(--text); }
      .actions { display: flex; gap: 0.75rem; flex-wrap: wrap; margin-top: 1rem; }
      button { padding: 0.85rem 1rem; border: none; border-radius: 0.55rem; cursor: pointer; font-weight: 600; }
      button.save { background: #2563eb; color: white; }
      button.delete { background: #ef4444; color: white; }
      button.add { background: #10b981; color: white; }
      .summary { margin-top: 1rem; padding: 1rem; background: var(--summary-bg); border-radius: 0.75rem; border: 1px solid var(--border); }
      .summary strong { font-size: 1.15rem; }
      .footer { margin-top: 1.3rem; color: var(--text-muted); font-size: 0.95rem; }
      .menu-toggle { display:inline-flex; align-items:center; justify-content:center; width:48px; height:48px; border:1px solid var(--border); border-radius:999px; background:var(--surface); color:var(--text); cursor:pointer; box-shadow: 0 2px 6px rgba(15,23,42,0.08); font-size:1.1rem; }
      .nav-menu { display:none; position:absolute; top:62px; left:0; width:220px; background:var(--surface); border:1px solid var(--border); border-radius:0.85rem; box-shadow:0 16px 32px rgba(15,23,42,0.12); z-index:20; }
      [data-theme="dark"] .nav-menu { box-shadow:0 16px 32px rgba(0,0,0,0.6); }
      .nav-menu.open { display:block; }
      .nav-menu a { display:block; padding:0.9rem 1rem; color:var(--text); text-decoration:none; border-bottom:1px solid var(--border); font-weight:700; }
      .nav-menu a:last-child { border-bottom:none; }
      .nav-menu a.active { color:var(--nav-active-color); background:var(--nav-active-bg); }
      .privacy-toggle { display:inline-flex; align-items:center; justify-content:center; width:48px; height:48px; border-radius:999px; border:1px solid var(--border); background:var(--surface); color:var(--text); cursor:pointer; font-weight:700; box-shadow: 0 2px 6px rgba(15,23,42,0.08); font-size:1.1rem; line-height:1; }
      .privacy-toggle:hover { background: var(--surface-2); }
      body.privacy-mode .sensitive { filter: blur(8px); user-select: none; }
      body.privacy-mode .sensitive * { pointer-events: none; }
      body.privacy-mode .sensitive-input { color: transparent !important; text-shadow: 0 0 10px currentColor; caret-color: transparent; }
      body.privacy-mode .sensitive-input::placeholder { color: transparent !important; }
      body.privacy-mode .chart-card, body.privacy-mode canvas { display:none !important; }
      .nav-button { padding:0.75rem 1rem; border-radius:0.75rem; border:1px solid var(--border); background:var(--surface); color:var(--text); text-decoration:none; font-weight:700; }
      .nav-button.active { background:#2563eb; color:#fff; border-color:transparent; }
      .chart-card { margin-top: 1.5rem; padding: 1rem; background: var(--surface-2); border-radius: 1rem; border: 1px solid var(--border); }
      .chart-legend-item { display:flex;align-items:center;gap:10px;padding:0.5rem 0.25rem;border-radius:8px;background:var(--surface);border:1px solid var(--border);color:var(--text); }
      .chart-legend-color { width:14px;height:14px;border-radius:4px;flex-shrink:0; }
      .settings-card { margin-top: 1.25rem; padding: 1rem; background: var(--surface-2); border-radius: 1rem; border: 1px solid var(--border); }
      .settings-form { display:grid; gap:1rem; }
      .settings-row { display:grid; grid-template-columns: 1fr auto; gap: 1rem; align-items:center; padding: 0.9rem; background:var(--surface); border-radius:0.9rem; border:1px solid var(--border); }
      .settings-account-row { position: relative; overflow: hidden; border-radius: 0.9rem; background: var(--surface); border: 1px solid var(--border); margin-bottom: 0.75rem; }
      .settings-account-inner { display:grid; grid-template-columns: 2fr 1fr; gap: 0.75rem; align-items:center; padding: 0.85rem 1rem; transition: transform 0.2s ease; }
      .settings-account-details { display:flex; flex-direction:column; gap:0.25rem; }
      .account-hover-wrap { position: relative; }
      .account-hover-card { position:absolute; left:0; top:calc(100% + 8px); min-width:260px; background:var(--surface); border:1px solid var(--border); border-radius:0.85rem; box-shadow:0 16px 32px rgba(15,23,42,0.18); padding:0.85rem 0.95rem; z-index:25; opacity:0; visibility:hidden; transform:translateY(-6px); transition: opacity 0.15s ease, transform 0.15s ease, visibility 0.15s ease; }
      .account-hover-wrap:hover .account-hover-card { opacity:1; visibility:visible; transform:translateY(0); }
      .account-hover-card .hover-label { font-size:0.72rem; text-transform:uppercase; letter-spacing:0.05em; color:var(--text-muted); font-weight:700; }
      .account-hover-card .hover-value { margin-top:0.15rem; font-weight:700; color:var(--text); }
      .account-name-link { font-weight:600; color:var(--account-link-color); text-decoration:none; }
      .account-name-link:hover { text-decoration:underline; }
      .settings-account-delete { position:absolute; top:0; right:0; bottom:0; width:84px; display:flex; align-items:center; justify-content:center; background:#ef4444; color:white; font-weight:700; border:none; transform: translateX(100%); transition: transform 0.2s ease; }
      .settings-account-row.swiped .settings-account-inner { transform: translateX(-84px); }
      .settings-account-row.swiped .settings-account-delete { transform: translateX(0); }
      input, select, textarea { background: var(--surface-2); color: var(--text); border: 1px solid var(--border); }
      input::placeholder, textarea::placeholder { color: var(--text-dim); }
      select option { background: var(--surface); color: var(--text); }
      .modal-box { background:var(--surface); border:1px solid var(--border); padding:18px; border-radius:12px; max-width:420px; margin:40px auto; width:calc(100% - 40px); }
      .modal-label { display:block; font-weight:600; margin-bottom:6px; color:var(--text); }
      .modal-input { width:100%; padding:8px; border-radius:6px; box-sizing:border-box; }
      .modal-select { padding:8px; border-radius:6px; width:100%; }
      .btn-cancel { padding:8px 10px; border-radius:6px; background:var(--cancel-bg); border:none; color:var(--cancel-text); cursor:pointer; font-weight:600; }
      .modal-display { padding:10px 12px; border:1px solid var(--border); border-radius:6px; background:var(--bg); color:var(--text); }

      @media (max-width: 700px) {
        body { margin: 0.6rem; }
        .card { padding: 0.75rem; border-radius: 0; box-shadow: none; max-width: 100%; }
        .account-row { grid-template-columns: 1fr !important; gap: 0.25rem; padding: 0.4rem; margin-bottom:0.25rem; }
        .account-row label, .account-row strong, .account-row div { font-size: 0.92rem; }
        .account-row > div { width: 100%; }
        .account-row > div:nth-child(2) { order: 2; text-align: left; margin-top: 3px; }
        .account-row > div:nth-child(3) { order: 3; display:flex; justify-content:flex-end; gap:6px; margin-top:3px }
        button, a.save { padding: 0.65rem 0.75rem; font-size: 14px; }
        .settings-account-inner { padding: 0.65rem 0.75rem; }
        .settings-account-row { margin-bottom: 0.5rem; }
        .actions { justify-content: space-between; }
        .card > div[style] { font-size: 15px; }
        #editModal > div { max-width: calc(100% - 20px); margin: 0; height: 100vh; border-radius: 0; display:flex; flex-direction:column; justify-content:flex-start; }
        #editModal form { flex: 1; display:flex; flex-direction:column; justify-content:space-between }
        #editModal form > div[style] { flex: none }
      }
    </style>
  </head>
  <body>
    <div class="card">
      <h1>Debt Tracker</h1>

      <div style="position:relative; margin-bottom:1rem; display:flex; align-items:center; gap:0.5rem;">
        <button id="menuToggle" class="menu-toggle" type="button" onclick="toggleMenu()" aria-expanded="false" aria-label="Open navigation menu">☰</button>
        <button id="privacyToggle" class="privacy-toggle" type="button" onclick="togglePrivacyMode()" aria-pressed="false" aria-label="Hide financials">👁</button>
        <div id="navMenu" class="nav-menu" aria-hidden="true">
          <a class="{% if page == 'tasks' %}active{% endif %}" href="{{ url_for('index') }}" onclick="closeMenu()">Tasks</a>
          <a class="{% if page == 'dashboard' %}active{% endif %}" href="{{ url_for('dashboard_page') }}" onclick="closeMenu()">Dashboard</a>
          <a class="{% if page == 'settings' %}active{% endif %}" href="{{ url_for('settings_page') }}" onclick="closeMenu()">Settings</a>
          <a class="{% if page == 'history' %}active{% endif %}" href="{{ url_for('history_page') }}" onclick="closeMenu()">History</a>
        </div>
      </div>

      {% if page == 'tasks' %}
        <div style="display:grid;grid-template-columns:1fr 1fr;gap:0.75rem;margin-top:1rem;">
          <div style="background:var(--surface);border:1px solid var(--border);border-radius:1rem;padding:1.1rem 1.25rem;">
            <div style="font-size:0.72rem;font-weight:700;color:var(--text-muted);text-transform:uppercase;letter-spacing:0.06em;margin-bottom:0.35rem;">Outstanding</div>
            <div class="sensitive" style="font-size:1.7rem;font-weight:800;color:var(--text);line-height:1;">${{ "%.2f"|format(total) }}</div>
          </div>
          <div style="background:var(--surface);border:1px solid var(--border);border-radius:1rem;padding:1.1rem 1.25rem;">
            <div style="font-size:0.72rem;font-weight:700;color:var(--text-muted);text-transform:uppercase;letter-spacing:0.06em;margin-bottom:0.35rem;">Left this month</div>
            <div class="sensitive" style="font-size:1.7rem;font-weight:800;line-height:1;color:{{ '#16a34a' if remaining_this_month == 0 else 'var(--text)' }};">${{ "%.2f"|format(remaining_this_month) }}</div>
          </div>
        </div>

        {% if view_mode == 'chart' %}
          <div class="chart-card">
            <h2 style="margin-top:0;">Debt distribution</h2>
            <canvas id="debtChart" width="340" height="340" style="max-width:100%;display:block;margin:0 auto;border-radius:16px;background:var(--surface)"></canvas>
            <div id="chartLegend" style="margin-top:16px;display:grid;grid-template-columns:1fr 1fr;gap:8px"></div>
          </div>
          <div class="chart-card" style="margin-top:1rem;">
            <h2 style="margin-top:0;">Outstanding balance trend</h2>
            <canvas id="trendChart" width="680" height="260" style="max-width:100%;display:block;margin:0 auto;border-radius:16px;background:var(--surface)"></canvas>
          </div>
        {% elif view_mode == 'list' %}
          <div style="display:flex;align-items:center;justify-content:space-between;margin-top:1.5rem;flex-wrap:wrap;gap:8px;">
            <h2 style="margin:0;display:flex;align-items:baseline;gap:10px;">Accounts <span style="font-size:0.85rem;font-weight:400;color:var(--text-muted);">{{ current_month_label }}</span></h2>
            <a href="{{ url_for('index', sort=sort_by, order=order, view=view_mode, hide_paid='0' if hide_paid else '1') }}" style="font-size:0.82rem;font-weight:600;padding:4px 12px;border-radius:999px;border:1px solid var(--border);background:var(--surface-2);color:var(--text-muted);text-decoration:none;">{{ 'Show paid' if hide_paid else 'Hide paid' }}</a>
          </div>

          {% set ns = namespace(monthly_total=0) %}
          {% for account in accounts %}{% if account.category == 'recurring' %}{% set ns.monthly_total = ns.monthly_total + (account.recurring_amount or 0) %}{% elif account.min_payment %}{% set ns.monthly_total = ns.monthly_total + account.min_payment %}{% endif %}{% endfor %}
          <div style="font-size:0.9rem;color:var(--text-muted);margin-top:0.75rem;">Monthly obligations: <strong class="sensitive" style="color:var(--text);">${{ "%.2f"|format(ns.monthly_total) }}</strong></div>
          <div style="display:grid;grid-template-columns:2fr 1fr 1fr 1fr 1fr;gap:0.75rem;font-weight:700;margin-top:12px;margin-bottom:6px;align-items:center;color:var(--text-muted);">
            <div role="button" onclick="sortBy('name')" style="cursor:pointer;display:flex;align-items:center;gap:8px;">Name {% if sort_by == 'name' %}<span style="font-size:0.85rem;color:var(--text-dim);">{{ '↓' if order == 'desc' else '↑' }}</span>{% endif %}</div>
            <div role="button" onclick="sortBy('type')" style="cursor:pointer;display:flex;align-items:center;gap:8px;">Type {% if sort_by == 'type' %}<span style="font-size:0.85rem;color:var(--text-dim);">{{ '↓' if order == 'desc' else '↑' }}</span>{% endif %}</div>
            <div role="button" onclick="sortBy('due_date')" style="cursor:pointer;display:flex;align-items:center;gap:8px;">Due {% if sort_by == 'due_date' %}<span style="font-size:0.85rem;color:var(--text-dim);">{{ '↓' if order == 'desc' else '↑' }}</span>{% endif %}</div>
            <div role="button" onclick="sortBy('amount')" style="cursor:pointer;text-align:right;display:flex;justify-content:flex-end;align-items:center;gap:8px;">{% if sort_by == 'amount' %}<span style="font-size:0.85rem;color:var(--text-dim);">{{ '↓' if order == 'desc' else '↑' }}</span>{% endif %}Outstanding</div>
            <div style="text-align:right;font-weight:700;">Monthly</div>
          </div>
          <form method="post">
            {% for account in accounts %}
            <div class="account-row month-account-row" role="button" tabindex="0" data-idx="{{ account.orig_idx }}" data-account-name="{{ account.name|e }}" data-account-type="{{ account.type or 'Other' }}" data-account-category="{{ account.category or 'debt' }}" data-balance="{{ account.balance|float }}" data-min-payment="{{ account.min_payment|default('', true) }}" data-recurring-amount="{{ account.recurring_amount|default('', true) }}" data-recurring-frequency="{{ account.recurring_frequency or '' }}" data-paid="{{ 'true' if account.paid_this_month else 'false' }}" data-paid1="{{ 'true' if account.paid_1_this_month else 'false' }}" data-paid2="{{ 'true' if account.paid_2_this_month else 'false' }}" data-date1="{{ account.due_date or '' }}" data-date2="{{ account.due_date_2 or '' }}" data-last-updated="{{ account.last_updated or '' }}" data-month-status="{{ account.paid_status_by_month|default({}, true)|tojson }}" style="border-left:6px solid {{ type_colors.get(account.type, '#64748b') }};grid-template-columns:2fr 1fr 1fr 1fr 1fr;cursor:pointer;">
              <div>
                <div style="display:flex;align-items:center;gap:8px;flex-wrap:wrap;">
                  <strong>{{ account.name }}</strong>
                  {% if account.category == 'recurring' %}<span class="sensitive" style="font-size:0.72rem;font-weight:700;padding:2px 7px;border-radius:999px;background:var(--surface);border:1px solid var(--border);color:var(--text-muted);">Recurring</span>{% endif %}
                  <span class="month-status-badge" data-category="{{ account.category }}" data-semi="{{ 'true' if account.recurring_frequency == 'semi-monthly' else 'false' }}" data-paid1-current="{{ 'true' if account.paid_1_this_month else 'false' }}" data-paid2-current="{{ 'true' if account.paid_2_this_month else 'false' }}" data-balance="{{ account.balance|float }}" style="font-size:0.72rem;font-weight:700;padding:2px 7px;border-radius:999px;letter-spacing:0.03em;"></span>
                </div>
                {% if account.interest_rate %}
                  <div style="font-size:0.85rem;color:var(--text-muted);margin-top:3px;">{{ account.interest_rate }}% APR</div>
                {% endif %}
              </div>
              <div style="font-weight:600;color:var(--text-muted);">{{ account.type or 'Other' }}</div>
              <div style="color:var(--text-muted);">
                {% if account.days_until is not none %}
                  {% if account.days_until == 0 %}
                    {% set badge_bg = '#dc2626' %}{% set badge_text = 'white' %}{% set due_label = 'Today' %}
                  {% elif account.days_until <= 3 %}
                    {% set badge_bg = '#dc2626' %}{% set badge_text = 'white' %}{% set due_label = 'In ' ~ account.days_until ~ ' day' ~ ('s' if account.days_until != 1 else '') %}
                  {% elif account.days_until <= 7 %}
                    {% set badge_bg = '#f59e0b' %}{% set badge_text = 'white' %}{% set due_label = 'In ' ~ account.days_until ~ ' days' %}
                  {% else %}
                    {% set badge_bg = 'var(--surface)' %}{% set badge_text = 'var(--text-muted)' %}{% set due_label = 'In ' ~ account.days_until ~ ' days' %}
                  {% endif %}
                  <div style="font-size:0.8rem;font-weight:700;padding:2px 8px;border-radius:999px;background:{{ badge_bg }};color:{{ badge_text }};display:inline-block;">{{ due_label }}</div>
                  <div style="font-size:0.78rem;color:var(--text-dim);margin-top:2px;">
                    {% if account.due_dates_label %}{{ account.due_dates_label }}{% else %}{{ account.next_due_date }}{% endif %}
                  </div>
                {% else %}
                  {{ account.due_date or '—' }}
                {% endif %}
              </div>
              <div class="sensitive" style="font-size:0.78rem;font-weight:700;color:var(--text);text-align:right;align-self:start;">
                ${{ "%.2f"|format(account.balance|float) }}
              </div>
              <div class="sensitive" style="font-size:0.78rem;font-weight:700;color:var(--text);text-align:right;align-self:start;">
                ${{ "%.2f"|format((account.recurring_amount if account.category == 'recurring' else account.min_payment) or 0) }}
              </div>
            </div>
            {% endfor %}
          </form>
        {% endif %}
      {% elif page == 'dashboard' %}
        <div style="display:flex;align-items:center;justify-content:space-between;margin-top:1rem;flex-wrap:wrap;gap:8px;">
          <div>
            <h2 style="margin:0;">Dashboard</h2>
            <div style="font-size:0.9rem;color:var(--text-muted);margin-top:0.25rem;">High-level overview of your bills and payments</div>
          </div>
          <a class="nav-button" href="{{ url_for('index') }}">Open Tasks</a>
        </div>

        <div style="display:grid;grid-template-columns:repeat(3,1fr);gap:0.75rem;margin-top:1rem;">
          <div style="background:var(--surface);border:1px solid var(--border);border-radius:1rem;padding:1rem 1.1rem;">
            <div style="font-size:0.72rem;font-weight:700;color:var(--text-muted);text-transform:uppercase;letter-spacing:0.06em;margin-bottom:0.35rem;">Current month due</div>
            <div class="sensitive" style="font-size:1.55rem;font-weight:800;line-height:1;">${{ "%.2f"|format(remaining_this_month) }}</div>
          </div>
          <div style="background:var(--surface);border:1px solid var(--border);border-radius:1rem;padding:1rem 1.1rem;">
            <div style="font-size:0.72rem;font-weight:700;color:var(--text-muted);text-transform:uppercase;letter-spacing:0.06em;margin-bottom:0.35rem;">Debt total</div>
            <div class="sensitive" style="font-size:1.55rem;font-weight:800;line-height:1;">${{ "%.2f"|format(total) }}</div>
          </div>
          <div style="background:var(--surface);border:1px solid var(--border);border-radius:1rem;padding:1rem 1.1rem;">
            <div style="font-size:0.72rem;font-weight:700;color:var(--text-muted);text-transform:uppercase;letter-spacing:0.06em;margin-bottom:0.35rem;">Recurring monthly</div>
            <div class="sensitive" style="font-size:1.55rem;font-weight:800;line-height:1;">${{ "%.2f"|format(monthly_recurring) }}</div>
          </div>
        </div>

        <div style="display:grid;grid-template-columns:1.1fr 0.9fr;gap:0.75rem;margin-top:1rem;">
          <div class="chart-card" style="margin-top:0;">
            <h2 style="margin-top:0;">Debt breakdown</h2>
            <canvas id="debtChart" width="340" height="340" style="max-width:100%;display:block;margin:0 auto;border-radius:16px;background:var(--surface)"></canvas>
          </div>
          <div class="chart-card" style="margin-top:0;">
            <h2 style="margin-top:0;">Snapshot</h2>
            <div style="display:grid;gap:0.75rem;">
              <div style="padding:0.9rem;border-radius:0.85rem;border:1px solid var(--border);background:var(--surface);">
                <div style="font-size:0.78rem;color:var(--text-muted);text-transform:uppercase;font-weight:700;">Month</div>
                <div style="font-size:1.15rem;font-weight:800;margin-top:0.25rem;">{{ current_month_label }}</div>
              </div>
              <div style="padding:0.9rem;border-radius:0.85rem;border:1px solid var(--border);background:var(--surface);">
                <div style="font-size:0.78rem;color:var(--text-muted);text-transform:uppercase;font-weight:700;">Open items</div>
                <div style="font-size:1.15rem;font-weight:800;margin-top:0.25rem;">{{ accounts|length }}</div>
              </div>
              <div style="padding:0.9rem;border-radius:0.85rem;border:1px solid var(--border);background:var(--surface);">
                <div style="font-size:0.78rem;color:var(--text-muted);text-transform:uppercase;font-weight:700;">Remaining this month</div>
                <div class="sensitive" style="font-size:1.15rem;font-weight:800;margin-top:0.25rem;">${{ "%.2f"|format(remaining_this_month) }}</div>
              </div>
            </div>
          </div>
        </div>

        <div class="chart-card" style="margin-top:1rem;">
          <h2 style="margin-top:0;">Trend over time</h2>
          <canvas id="trendChart" width="680" height="260" style="max-width:100%;display:block;margin:0 auto;border-radius:16px;background:var(--surface)"></canvas>
        </div>
      {% elif page == 'settings' %}
        <div class="settings-card">
          <h2>Settings</h2>
          <form method="post" class="settings-form">
            <input type="hidden" name="action" value="update_settings" />
            <label style="display:flex;align-items:center;gap:10px;">
              <input type="checkbox" name="show_zero" value="1" {% if show_zero %}checked{% endif %}/> Show zero balances on Dashboard
            </label>
            <div style="display:flex;align-items:center;justify-content:space-between;">
              <span style="font-weight:600;">Theme</span>
              <button type="button" onclick="toggleTheme()" style="padding:6px 18px;border-radius:999px;border:1px solid var(--border);background:var(--surface-2);color:var(--text);cursor:pointer;font-weight:600;font-size:0.9rem;" id="themeToggle"></button>
            </div>
            <button class="save" type="submit">Save settings</button>
          </form>
          {% if status %}
            <div style="padding:0.85rem 1rem;margin-top:0.75rem;border-radius:0.85rem;background:var(--success-bg);color:var(--success-text);font-weight:600;">{{ status }}</div>
          {% endif %}

          <h2 style="margin-top:1.5rem;">Export &amp; Import</h2>
          <div style="display:flex;flex-direction:column;gap:0.75rem;">
            <div>
              <p style="margin:0 0 6px;color:var(--text-muted);font-size:0.9rem;">Download a full save file you can share or restore later.</p>
              <a href="{{ url_for('export_data') }}" style="display:inline-flex;align-items:center;justify-content:center;padding:0.75rem 1rem;border-radius:0.55rem;background:#2563eb;color:white;text-decoration:none;font-weight:600;">Export save file</a>
            </div>
            <div>
              <p style="margin:0 0 6px;color:var(--text-muted);font-size:0.9rem;">Download an Excel workbook with charts and a dashboard that can be edited and imported back into the app.</p>
              <a href="{{ url_for('export_excel') }}" style="display:inline-flex;align-items:center;justify-content:center;padding:0.75rem 1rem;border-radius:0.55rem;background:#16a34a;color:white;text-decoration:none;font-weight:600;">Export Excel workbook</a>
            </div>
            <div>
              <p style="margin:0 0 6px;color:var(--text-muted);font-size:0.9rem;">Restore from a previously exported save file. <strong style="color:var(--text);">This replaces all current data.</strong></p>
              <form method="post" action="{{ url_for('import_data') }}" enctype="multipart/form-data" style="display:flex;align-items:center;gap:0.5rem;flex-wrap:wrap;">
                <input type="file" name="save_file" accept=".json" required style="font-size:0.9rem;color:var(--text);background:var(--surface-2);border:1px solid var(--border);border-radius:0.5rem;padding:6px 10px;cursor:pointer;" />
                <button type="submit" class="save" style="padding:0.65rem 1rem;" onclick="return confirm('This will replace all current data with the imported file. Continue?')">Import</button>
              </form>
            </div>
            <div>
              <p style="margin:0 0 6px;color:var(--text-muted);font-size:0.9rem;">Import changes from an Excel workbook. The `Accounts`, `History`, `Monthly Totals`, and `Settings` sheets are read back in.</p>
              <form method="post" action="{{ url_for('import_excel') }}" enctype="multipart/form-data" style="display:flex;align-items:center;gap:0.5rem;flex-wrap:wrap;">
                <input type="file" name="excel_file" accept=".xlsx" required style="font-size:0.9rem;color:var(--text);background:var(--surface-2);border:1px solid var(--border);border-radius:0.5rem;padding:6px 10px;cursor:pointer;" />
                <button type="submit" class="save" style="padding:0.65rem 1rem;" onclick="return confirm('This will replace all current data with the imported workbook. Continue?')">Import Excel</button>
              </form>
            </div>
          </div>

          <h2 style="margin-top:1.5rem;">History CSV</h2>
          <div class="settings-form" style="padding:0;">
            <a href="{{ url_for('download_history') }}" style="display:inline-flex;align-items:center;justify-content:center;padding:0.85rem 1rem;border-radius:0.55rem;background:#2563eb;color:white;text-decoration:none;font-weight:600;">Download history CSV</a>
          </div>

          <div style="margin-top:1.5rem;">
            <button type="button" class="add" onclick="document.getElementById('addAccountModal').style.display='flex'">+ Add account</button>
          </div>

          <h2 style="margin-top:1.5rem;">Manage accounts</h2>
          <div id="settingsAccountList">
            {% for account in accounts %}
              <div class="settings-account-row" data-idx="{{ account.orig_idx }}">
                <div class="settings-account-inner">
                  <div class="settings-account-details">
                    <div class="account-hover-wrap">
                      <a href="{{ url_for('account_page', idx=account.orig_idx) }}" class="account-name-link">{{ account.name }}</a>
                      <div class="account-hover-card">
                        <div>
                          <div class="hover-label">Last updated</div>
                          <div class="hover-value">{{ account.last_updated[:10] if account.last_updated else 'No update yet' }}</div>
                        </div>
                        <div style="margin-top:0.75rem;">
                          <div class="hover-label">Minimum monthly payment</div>
                          <div class="hover-value">{% if account.min_payment %}<span class="sensitive">${{ "%.2f"|format(account.min_payment|float) }}</span>{% else %}—{% endif %}</div>
                        </div>
                        <div style="margin-top:0.75rem;">
                          <div class="hover-label">Outstanding balance</div>
                          <div class="hover-value">{% if account.category == 'recurring' %}<span class="sensitive">${{ "%.2f"|format(account.recurring_amount or 0) }}</span>{% else %}<span class="sensitive">${{ "%.2f"|format(account.balance) }}</span>{% endif %}</div>
                        </div>
                      </div>
                    </div>
                    <span style="font-size:0.9rem;color:var(--text-muted);">{{ account.type or 'Other' }}{% if account.owner %} · {{ account.owner }}{% endif %}</span>
                    <div style="font-size:0.85rem;color:var(--text-muted);">{% if account.due_date %}Due {{ account.due_date }}{% endif %}{% if account.interest_rate %} · {{ account.interest_rate }}% APR{% endif %}</div>
                  </div>
                  <div style="display:flex;flex-direction:column;align-items:flex-end;gap:6px;">
                    <button type="button" onclick="archiveAccount({{ account.orig_idx }})" style="font-size:0.75rem;padding:2px 8px;border-radius:999px;border:1px solid var(--border);background:var(--surface-2);color:var(--text-muted);cursor:pointer;">Archive</button>
                  </div>
                </div>
                <button class="settings-account-delete" type="button">Delete</button>
              </div>
            {% endfor %}
          </div>
          <p style="margin-top:0.75rem;color:var(--text-muted);font-size:0.95rem;">Swipe left on an account to reveal the delete button on touch devices.</p>

          {% if archived_accounts %}
          <h2 style="margin-top:1.75rem;color:var(--text-muted);">Archived</h2>
          <div id="archivedAccountList">
            {% for account in archived_accounts %}
              <div class="settings-account-row" data-idx="{{ account.orig_idx }}" style="opacity:0.55;">
                <div class="settings-account-inner">
                  <div class="settings-account-details">
                    <div class="account-hover-wrap">
                      <span class="account-name-link" style="cursor:default;">{{ account.name }}</span>
                      <div class="account-hover-card">
                        <div>
                          <div class="hover-label">Last updated</div>
                          <div class="hover-value">{{ account.last_updated[:10] if account.last_updated else 'No update yet' }}</div>
                        </div>
                        <div style="margin-top:0.75rem;">
                          <div class="hover-label">Minimum monthly payment</div>
                          <div class="hover-value">{% if account.min_payment %}<span class="sensitive">${{ "%.2f"|format(account.min_payment|float) }}</span>{% else %}—{% endif %}</div>
                        </div>
                        <div style="margin-top:0.75rem;">
                          <div class="hover-label">Outstanding balance</div>
                          <div class="hover-value">{% if account.category == 'recurring' %}<span class="sensitive">${{ "%.2f"|format(account.recurring_amount or 0) }}</span>{% else %}<span class="sensitive">${{ "%.2f"|format(account.balance) }}</span>{% endif %}</div>
                        </div>
                      </div>
                    </div>
                    <span style="font-size:0.9rem;color:var(--text-muted);">{{ account.type or 'Other' }}{% if account.owner %} · {{ account.owner }}{% endif %}</span>
                  </div>
                  <div style="display:flex;flex-direction:column;align-items:flex-end;gap:6px;">
                    {% if account.category == 'recurring' %}
                      <div class="sensitive" style="font-weight:600;">${{ "%.2f" | format(account.recurring_amount or 0) }}</div>
                    {% else %}
                      <div class="sensitive" style="font-weight:600;">${{ "%.2f" | format(account.balance) }}</div>
                    {% endif %}
                    <button type="button" onclick="unarchiveAccount({{ account.orig_idx }})" style="font-size:0.75rem;padding:2px 8px;border-radius:999px;border:1px solid var(--border);background:var(--surface-2);color:var(--text-muted);cursor:pointer;">Unarchive</button>
                  </div>
                </div>
                <button class="settings-account-delete" type="button">Delete</button>
              </div>
            {% endfor %}
          </div>
          {% endif %}
        </div>
      {% elif page == 'account_detail' %}
        <div class="settings-card">
          <h2>{{ account.name }}</h2>
          {% if account.category == 'recurring' %}
            <p class="sensitive" style="margin-top:0.25rem;color:var(--text-muted);">${{ "%.2f"|format(account.recurring_amount or 0) }} / {% if account.recurring_frequency == 'semi-monthly' %}semi-monthly (2x / mo){% elif account.recurring_frequency %}{{ account.recurring_frequency }}{% else %}monthly{% endif %}</p>
          {% else %}
            <p class="sensitive" style="margin-top:0.25rem;color:var(--text-muted);">Current balance: ${{ "%.2f"|format(account.balance) }}</p>
          {% endif %}

          <h3 style="margin-top:1.25rem;margin-bottom:0.75rem;">Account details</h3>
          <form method="post" action="{{ url_for('edit_account', idx=idx) }}" class="settings-form" style="margin-bottom:1.5rem;">
            <input type="hidden" name="category" value="{{ account.category or 'debt' }}" />
            <input type="hidden" name="return_to" value="{{ return_to }}" />
            <div style="display:grid;grid-template-columns:1fr 1fr;gap:0.75rem;">
              <div>
                <label style="font-weight:600;display:block;margin-bottom:4px;">Name</label>
                <input type="text" name="name" value="{{ account.name }}" required style="width:100%;padding:8px;border-radius:6px;box-sizing:border-box;" />
              </div>
              <div>
                <label style="font-weight:600;display:block;margin-bottom:4px;">Owner</label>
                <input type="text" name="owner" value="{{ account.owner or '' }}" placeholder="e.g. John, Jane" style="width:100%;padding:8px;border-radius:6px;box-sizing:border-box;" />
              </div>
            </div>
            {% if account.category == 'recurring' %}
            <div style="display:grid;grid-template-columns:1fr 1fr;gap:0.75rem;">
              <div>
                <label style="font-weight:600;display:block;margin-bottom:4px;">Amount</label>
                <input type="number" name="recurring_amount" step="0.01" min="0" value="{{ account.recurring_amount or '' }}" placeholder="0.00" style="width:100%;padding:8px;border-radius:6px;box-sizing:border-box;" />
              </div>
              <div>
                <label style="font-weight:600;display:block;margin-bottom:4px;">Frequency</label>
                <select name="recurring_frequency" style="width:100%;padding:8px;border-radius:6px;" onchange="toggleDetailFrequency(this.value)">
                  {% set freq = account.recurring_frequency or 'monthly' %}
                  <option value="monthly" {% if freq == 'monthly' %}selected{% endif %}>Monthly</option>
                  <option value="semi-monthly" {% if freq == 'semi-monthly' %}selected{% endif %}>Semi-monthly (2x / mo)</option>
                  <option value="yearly" {% if freq == 'yearly' %}selected{% endif %}>Yearly</option>
                  <option value="quarterly" {% if freq == 'quarterly' %}selected{% endif %}>Quarterly</option>
                  <option value="weekly" {% if freq == 'weekly' %}selected{% endif %}>Weekly</option>
                </select>
              </div>
            </div>
            <div style="display:grid;grid-template-columns:1fr 1fr;gap:0.75rem;">
              <div>
                <label style="font-weight:600;display:block;margin-bottom:4px;">Type</label>
                <select name="type" style="width:100%;padding:8px;border-radius:6px;">
                  {% for t in ['Subscription','Streaming','Utility','Insurance','Other'] %}
                    <option value="{{ t }}" {% if (account.type or 'Other') == t %}selected{% endif %}>{{ t }}</option>
                  {% endfor %}
                </select>
              </div>
            </div>
            {% else %}
            <div style="display:grid;grid-template-columns:1fr 1fr;gap:0.75rem;">
              <div>
                <label style="font-weight:600;display:block;margin-bottom:4px;">Type</label>
                <select name="type" style="width:100%;padding:8px;border-radius:6px;">
                  {% for t in ['Credit Card','Car Loan','Mortgage','Student Loan','Personal Loan','Other'] %}
                    <option value="{{ t }}" {% if (account.type or 'Other') == t %}selected{% endif %}>{{ t }}</option>
                  {% endfor %}
                </select>
              </div>
              <div>
                <label style="font-weight:600;display:block;margin-bottom:4px;">Interest rate (%)</label>
                <input type="number" name="interest_rate" step="0.01" min="0" value="{{ account.interest_rate or '' }}" placeholder="%" style="width:100%;padding:8px;border-radius:6px;box-sizing:border-box;" />
              </div>
            </div>
            <div style="display:grid;grid-template-columns:1fr 1fr;gap:0.75rem;">
              <div>
                <label style="font-weight:600;display:block;margin-bottom:4px;">Outstanding amount</label>
                <input type="number" class="sensitive-input" name="balance" step="0.01" min="0" value="{{ '%.2f'|format(account.balance|float) }}" placeholder="0.00" style="width:100%;padding:8px;border-radius:6px;box-sizing:border-box;" />
              </div>
              <div>
                <label style="font-weight:600;display:block;margin-bottom:4px;">Payment Due</label>
                <input type="number" name="min_payment" step="0.01" min="0" value="{{ account.min_payment or '' }}" placeholder="0.00" style="width:100%;padding:8px;border-radius:6px;box-sizing:border-box;" />
              </div>
            </div>
            {% endif %}
            {% set is_semi = account.recurring_frequency == 'semi-monthly' %}
            <div id="detail_due_single" style="display:{% if is_semi %}none{% else %}block{% endif %};">
              <label style="font-weight:600;display:block;margin-bottom:4px;">Due day</label>
              <select id="detail_due_date_select" name="due_date_type" style="width:100%;padding:8px;border-radius:6px;">
                <option value="1st" {% if account.due_date == '1st' %}selected{% endif %}>1st</option>
                <option value="15th" {% if account.due_date == '15th' %}selected{% endif %}>15th</option>
                <option value="last" {% if account.due_date == 'last' %}selected{% endif %}>Last day</option>
                <option value="custom" {% if account.due_date and account.due_date not in ('1st','15th','last','unknown') %}selected{% endif %}>Custom day</option>
                <option value="unknown" {% if account.due_date == 'unknown' %}selected{% endif %}>Unknown</option>
              </select>
              <select id="detail_due_date_custom" name="due_date" style="display:{% if account.due_date and account.due_date not in ('1st','15th','last','unknown') %}block{% else %}none{% endif %};width:100%;padding:8px;border-radius:6px;margin-top:6px;">
                {% for d in range(1, 32) %}<option value="{{ d }}" {% if account.due_date == d|string %}selected{% endif %}>{{ d }}</option>{% endfor %}
              </select>
            </div>
            <div id="detail_due_semi" style="display:{% if is_semi %}block{% else %}none{% endif %};">
              <label style="font-weight:600;display:block;margin-bottom:4px;">Billing dates</label>
              <div style="display:grid;grid-template-columns:1fr 1fr;gap:0.75rem;">
                <div>
                  <label style="font-size:0.8rem;color:var(--text-muted);display:block;margin-bottom:4px;">First day</label>
                  <select name="semi_date_1" style="width:100%;padding:8px;border-radius:6px;">
                    {% for d in range(1, 32) %}<option value="{{ d }}" {% if account.due_date == d|string %}selected{% endif %}>{{ d }}</option>{% endfor %}
                  </select>
                </div>
                <div>
                  <label style="font-size:0.8rem;color:var(--text-muted);display:block;margin-bottom:4px;">Second day</label>
                  <select name="semi_date_2" style="width:100%;padding:8px;border-radius:6px;">
                    {% for d in range(1, 32) %}<option value="{{ d }}" {% if account.due_date_2 == d|string %}selected{% endif %}>{{ d }}</option>{% endfor %}
                  </select>
                </div>
              </div>
            </div>
            <div style="display:flex;gap:0.75rem;flex-wrap:wrap;align-items:center;">
              <button class="save" type="submit">Save changes</button>
              <button type="button" class="delete" onclick="if(confirm('Delete this account? This cannot be undone.')) fetch('/delete/{{ idx }}',{method:'POST'}).then(r=>{ if(r.ok) window.location='/settings'; })">Delete account</button>
            </div>
          </form>

          <h3 style="margin-top:0;margin-bottom:0.75rem;">Monthly totals</h3>
          <form method="post" action="{{ url_for('update_month', idx=idx) }}" class="settings-form" style="margin-bottom:1rem;">
            <label style="font-weight:600;">Enter month and total on each line</label>
            <textarea name="monthly_totals" placeholder="2024-05 1500.00\n2024-06 1425.50\n..." style="width:100%;padding:10px;border-radius:8px;min-height:140px;box-sizing:border-box;"></textarea>
            <button class="save" type="submit">Save monthly totals</button>
          </form>
          {% if monthly_totals|length > 0 %}
            <div style="display:grid;gap:0.75rem;">
              {% for entry in monthly_totals %}
                <div style="display:grid;grid-template-columns:1fr 1fr auto;gap:0.75rem;align-items:center;padding:0.85rem 0.95rem;border-radius:0.9rem;background:var(--surface);border:1px solid var(--border);">
                  <div>{{ entry.month }}</div>
                  <div style="font-weight:700;">${{ "%.2f"|format(entry.total) }}</div>
                  <button type="button" class="save" style="padding:0.7rem 1rem;" onclick="openMonthEditor('{{ entry.month }}', '{{ entry.total|float }}')">Edit</button>
                </div>
              {% endfor %}
            </div>
          {% else %}
            <p style="color:var(--text-muted);">No monthly totals available yet for this account.</p>
          {% endif %}
          <div id="monthlyEditSection" style="display:none;margin-top:1rem;padding:1rem;border-radius:0.9rem;background:var(--surface-2);border:1px solid var(--border);">
            <h4 style="margin-top:0;margin-bottom:0.75rem;">Edit monthly entry</h4>
            <form method="post" action="{{ url_for('update_month', idx=idx) }}" class="settings-form">
              <div style="display:grid;grid-template-columns:1fr 1fr;gap:0.75rem;">
                <div>
                  <label style="font-weight:600;">Month</label>
                  <input id="edit_month_input" name="month" type="month" required style="width:100%;padding:10px;border-radius:8px;box-sizing:border-box;" />
                </div>
                <div>
                  <label style="font-weight:600;">Total</label>
                  <input id="edit_total_input" name="total" type="number" step="0.01" min="0" required style="width:100%;padding:10px;border-radius:8px;box-sizing:border-box;" />
                </div>
              </div>
              <div style="display:flex;gap:0.75rem;flex-wrap:wrap;margin-top:0.75rem;">
                <button class="save" type="submit">Save entry</button>
                <button type="button" class="delete" onclick="hideMonthEditForm()">Cancel</button>
              </div>
            </form>
          </div>
          <div style="display:flex;gap:0.75rem;flex-wrap:wrap;margin-top:1rem;">
            <a class="nav-button" href="{{ url_for('settings_page') }}">Back to settings</a>
            <a class="nav-button" href="{{ url_for('index') }}">Back to dashboard</a>
          </div>
        </div>
      {% elif page == 'history' %}
        <div class="settings-card">
          <h2>Account history</h2>
          {% if history_entries|length == 0 %}
            <p style="color:var(--text-muted);">No history entries available yet.</p>
          {% else %}
            <div style="display:grid;gap:0.75rem; margin-top:1rem;">
              {% for event in history_entries %}
                <div style="padding:0.85rem 0.9rem;border-radius:0.9rem;background:var(--surface);border:1px solid var(--border);">
                  <div style="font-weight:700;">{{ event.action }}</div>
                  <div style="font-size:0.92rem;color:var(--text-muted);margin-top:0.15rem;">{{ event.timestamp }}</div>
                  <div style="font-size:0.95rem;color:var(--text);margin-top:0.5rem;font-weight:700;">{{ event.account_name }}</div>
                  <div style="font-size:0.92rem;color:var(--text-muted);margin-top:0.45rem;line-height:1.4;">
                    {% if event.prev_amount is not none %}
                      Amount: <span class="sensitive">${{ "%.2f" | format(event.amount) }}</span> (from <span class="sensitive">${{ "%.2f" | format(event.prev_amount) }}</span>)<br />
                    {% else %}
                      Amount: <span class="sensitive">${{ "%.2f" | format(event.amount) }}</span><br />
                    {% endif %}
                    Type: {{ event.new_type or event.type or 'Other' }}
                    {% if event.old_type and event.old_type != event.new_type %}(from {{ event.old_type }}){% endif %}
                    {% if event.old_name and event.old_name != event.new_name %}<br />Name: {{ event.old_name }} → {{ event.new_name }}{% endif %}
                    {% if event.new_due_date is not none or event.old_due_date is not none %}
                      <br />Due date: {{ event.new_due_date or 'none' }}{% if event.old_due_date and event.old_due_date != event.new_due_date %} (from {{ event.old_due_date }}){% endif %}
                    {% endif %}
                    {% if event.new_interest_rate is not none or event.old_interest_rate is not none %}
                      <br />Interest: {{ event.new_interest_rate if event.new_interest_rate is not none else '0' }}%{% if event.old_interest_rate and event.old_interest_rate != event.new_interest_rate %} (from {{ event.old_interest_rate }}%) {% endif %}
                    {% endif %}
                  </div>
                </div>
              {% endfor %}
            </div>
          {% endif %}
        </div>
      {% endif %}

      <div class="footer">
        <p>Debt values are stored locally in <code style="background:var(--code-bg);padding:2px 6px;border-radius:4px;font-size:0.9em;border:1px solid var(--border);">debt_tracker.db</code>.</p>
      </div>
    </div>

    <div id="addAccountModal" style="display:none;position:fixed;left:0;top:0;width:100%;height:100%;background:var(--modal-overlay);align-items:center;justify-content:center;z-index:50;" onclick="if(event.target===this)this.style.display='none'">
      <div class="modal-box">
        <h3 style="margin:0 0 14px;font-size:1.1rem;">Add account</h3>
        <form method="post" class="settings-form" style="gap:10px;">
          <input type="hidden" name="action" value="add_account" />
          <div>
            <label class="modal-label">Category</label>
            <select id="add_category" name="new_category" class="modal-select" style="border:1px solid var(--border);background:var(--bg);color:var(--text);" onchange="toggleAddCategory(this.value)">
              <option value="debt">Debt</option>
              <option value="recurring">Recurring expense</option>
            </select>
          </div>
          <div>
            <label class="modal-label">Account name</label>
            <input type="text" name="new_name" placeholder="e.g. Chase Sapphire" required class="modal-input" style="border:1px solid var(--border);background:var(--bg);color:var(--text);" />
          </div>
          <div>
            <label class="modal-label">Owner</label>
            <input type="text" name="new_owner" placeholder="e.g. John, Jane" class="modal-input" style="border:1px solid var(--border);background:var(--bg);color:var(--text);" />
          </div>
          <div id="add_debt_fields">
            <label class="modal-label">Balance</label>
            <input type="number" step="0.01" min="0" name="new_balance" placeholder="0.00" class="modal-input" style="border:1px solid var(--border);background:var(--bg);color:var(--text);" />
          </div>
          <div id="add_recurring_fields" style="display:none;gap:8px;flex-direction:column;">
            <div>
              <label class="modal-label">Amount</label>
              <input type="number" step="0.01" min="0" name="new_recurring_amount" placeholder="0.00" class="modal-input" style="border:1px solid var(--border);background:var(--bg);color:var(--text);" />
            </div>
            <div>
              <label class="modal-label">Frequency</label>
              <select name="new_recurring_frequency" class="modal-select" style="border:1px solid var(--border);background:var(--bg);color:var(--text);" onchange="toggleAddFrequency(this.value)">
                <option value="monthly">Monthly</option>
                <option value="semi-monthly">Semi-monthly (2x / mo)</option>
                <option value="yearly">Yearly</option>
                <option value="quarterly">Quarterly</option>
                <option value="weekly">Weekly</option>
              </select>
            </div>
          </div>
          <div id="add_due_date_single">
            <label class="modal-label">Billing date</label>
            <select id="add_due_date_select" name="new_due_date_type" class="modal-select" style="border:1px solid var(--border);background:var(--bg);color:var(--text);" onchange="document.getElementById('add_due_date_custom').style.display=this.value==='custom'?'block':'none'">
              <option value="1st">1st</option>
              <option value="15th">15th</option>
              <option value="last">Last day</option>
              <option value="custom">Custom day</option>
              <option value="unknown">Unknown</option>
            </select>
            <select id="add_due_date_custom" name="new_due_date" class="modal-select" style="display:none;margin-top:6px;border:1px solid var(--border);background:var(--bg);color:var(--text);">
              {% for d in range(1, 32) %}<option value="{{ d }}">{{ d }}</option>{% endfor %}
            </select>
          </div>
          <div id="add_due_date_semi" style="display:none;">
            <label class="modal-label">Billing dates</label>
            <div style="display:grid;grid-template-columns:1fr 1fr;gap:8px;">
              <div>
                <label style="font-size:0.8rem;color:var(--text-muted);display:block;margin-bottom:4px;">First day</label>
                <select name="new_semi_date_1" class="modal-select" style="border:1px solid var(--border);background:var(--bg);color:var(--text);">
                  {% for d in range(1, 32) %}<option value="{{ d }}">{{ d }}</option>{% endfor %}
                </select>
              </div>
              <div>
                <label style="font-size:0.8rem;color:var(--text-muted);display:block;margin-bottom:4px;">Second day</label>
                <select name="new_semi_date_2" class="modal-select" style="border:1px solid var(--border);background:var(--bg);color:var(--text);">
                  {% for d in range(1, 32) %}<option value="{{ d }}">{{ d }}</option>{% endfor %}
                </select>
              </div>
            </div>
          </div>
          <div id="add_interest_field">
            <label class="modal-label">Interest rate (%)</label>
            <input type="number" step="0.01" min="0" name="new_interest_rate" placeholder="e.g. 19.99" class="modal-input" style="border:1px solid var(--border);background:var(--bg);color:var(--text);" />
          </div>
          <div>
            <label class="modal-label">Type</label>
            <select id="add_type_debt" name="new_type" class="modal-select" style="border:1px solid var(--border);background:var(--bg);color:var(--text);">
              <option value="Credit Card">Credit Card</option>
              <option value="Car Loan">Car Loan</option>
              <option value="Mortgage">Mortgage</option>
              <option value="Student Loan">Student Loan</option>
              <option value="Personal Loan">Personal Loan</option>
              <option value="Other">Other</option>
            </select>
            <select id="add_type_recurring" name="new_type" class="modal-select" style="display:none;border:1px solid var(--border);background:var(--bg);color:var(--text);">
              <option value="Subscription">Subscription</option>
              <option value="Streaming">Streaming</option>
              <option value="Utility">Utility</option>
              <option value="Insurance">Insurance</option>
              <option value="Other">Other</option>
            </select>
          </div>
          <div style="display:flex;gap:8px;justify-content:flex-end;margin-top:4px;">
            <button type="button" onclick="document.getElementById('addAccountModal').style.display='none'" class="btn-cancel">Cancel</button>
            <button type="submit" class="add">Add account</button>
          </div>
        </form>
      </div>
    </div>

    <div id="updateModal" style="display:none;position:fixed;left:0;top:0;width:100%;height:100%;background:var(--modal-overlay);align-items:center;justify-content:center;z-index:50;">
      <div class="modal-box">
        <h3 id="updateModalTitle" style="margin-top:0;">Update balance</h3>
        <form id="updateForm" method="post" onsubmit="submitUpdate(event)">
          <div id="update_debt_fields">
            <div style="margin-bottom:8px;">
              <div class="modal-display" id="update_summary" style="display:grid;gap:0.45rem;">
                <div>
                  <div class="modal-label" style="margin-bottom:2px;">Account</div>
                  <div id="update_account_name" style="font-weight:700;">-</div>
                </div>
                <div style="display:grid;grid-template-columns:1fr 1fr;gap:0.75rem;">
                  <div>
                    <div class="modal-label" style="margin-bottom:2px;">Type</div>
                    <div id="update_account_type" style="font-weight:700;">-</div>
                  </div>
                  <div>
                    <div class="modal-label" style="margin-bottom:2px;">Last updated</div>
                    <div id="update_last_updated" style="font-weight:700;">No update yet</div>
                  </div>
                </div>
                <div style="display:grid;grid-template-columns:1fr 1fr;gap:0.75rem;">
                  <div>
                    <div class="modal-label" style="margin-bottom:2px;">Payment Due</div>
                    <div id="update_min_payment" style="font-weight:700;">—</div>
                  </div>
                  <div>
                    <div class="modal-label" style="margin-bottom:2px;">Outstanding balance</div>
                    <div id="update_current_balance" style="font-weight:700;">$0.00</div>
                  </div>
                </div>
              </div>
            </div>
            <div style="margin-bottom:8px;">
              <label class="modal-label" for="update_balance">Update balance</label>
              <input id="update_balance" name="new_balance" type="number" step="0.01" min="0" value="0.00" class="modal-input" />
            </div>
            <div style="margin-bottom:8px;">
              <label class="modal-label">Category</label>
              <select id="update_type" name="new_type" class="modal-select">
                <option value="Credit Card">Credit Card</option>
                <option value="Car Loan">Car Loan</option>
                <option value="Mortgage">Mortgage</option>
                <option value="Student Loan">Student Loan</option>
                <option value="Personal Loan">Personal Loan</option>
                <option value="Other">Other</option>
              </select>
            </div>
          </div>
          <div id="update_paid_single" style="margin-bottom:8px;">
            <label style="display:flex;align-items:center;gap:8px;cursor:pointer;">
              <input type="checkbox" id="update_mark_paid" style="width:16px;height:16px;accent-color:#16a34a;" />
              <span style="font-weight:600;color:var(--text);">Mark <span id="update_paid_month_label"></span> complete</span>
            </label>
          </div>
          <div id="update_paid_semi" style="display:none;margin-bottom:8px;display:none;flex-direction:column;gap:8px;">
            <label style="display:flex;align-items:center;gap:8px;cursor:pointer;">
              <input type="checkbox" id="update_mark_paid_1" style="width:16px;height:16px;accent-color:#16a34a;" />
              <span style="font-weight:600;color:var(--text);">Payment 1 (day <span id="update_date1_label"></span>) paid</span>
            </label>
            <label style="display:flex;align-items:center;gap:8px;cursor:pointer;">
              <input type="checkbox" id="update_mark_paid_2" style="width:16px;height:16px;accent-color:#16a34a;" />
              <span style="font-weight:600;color:var(--text);">Payment 2 (day <span id="update_date2_label"></span>) paid</span>
            </label>
          </div>
          <div id="updateResult" style="display:none;padding:10px;margin-bottom:8px;border-radius:6px;background:var(--success-bg);color:var(--success-text);font-weight:600;"></div>
          <div style="display:flex;gap:8px;justify-content:flex-end;flex-wrap:wrap">
            <a id="updateEditAccountLink" href="#" class="nav-button" style="text-decoration:none;">Edit account</a>
            <button type="button" onclick="closeUpdateModal()" class="btn-cancel">Cancel</button>
            <button type="submit" id="updateSubmitButton" style="padding:8px 10px;border-radius:6px;background:#2563eb;color:white;border:none;cursor:pointer;font-weight:600;">Save</button>
          </div>
        </form>
      </div>
    </div>

    <script>
      function setPrivacyMode(enabled) {
        document.body.classList.toggle('privacy-mode', enabled);
        const btn = document.getElementById('privacyToggle');
        if (btn) {
          btn.textContent = enabled ? '🙈' : '👁';
          btn.setAttribute('aria-label', enabled ? 'Show financials' : 'Hide financials');
          btn.setAttribute('aria-pressed', String(enabled));
        }
        localStorage.setItem('privacyMode', enabled ? '1' : '0');
      }

      function togglePrivacyMode() {
        setPrivacyMode(!(document.body.classList.contains('privacy-mode')));
      }

      (function() {
        setPrivacyMode(localStorage.getItem('privacyMode') === '1');
      })();

      // ── Theme ──────────────────────────────────────────────────────────
      function applyTheme(theme) {
        document.documentElement.setAttribute('data-theme', theme);
        const btn = document.getElementById('themeToggle');
        if (btn) btn.textContent = theme === 'dark' ? 'Light mode' : 'Dark mode';
        rerenderCharts();
      }
      function toggleTheme() {
        const current = document.documentElement.getAttribute('data-theme') || 'light';
        const next = current === 'dark' ? 'light' : 'dark';
        localStorage.setItem('theme', next);
        applyTheme(next);
      }
      (function() {
        const saved = localStorage.getItem('theme') || 'dark';
        applyTheme(saved);
      })();

      // ── CSS var helper for canvas rendering ───────────────────────────
      function cssVar(name) {
        return getComputedStyle(document.documentElement).getPropertyValue(name).trim();
      }

      // ── Nav ───────────────────────────────────────────────────────────
      function sortBy(field){
        const params = new URLSearchParams(window.location.search);
        const cur = params.get('sort');
        const curOrder = params.get('order') || 'desc';
        if(cur === field){
          params.set('order', curOrder === 'desc' ? 'asc' : 'desc');
        } else {
          params.set('sort', field);
          params.set('order', 'desc');
        }
        window.location.search = params.toString();
      }

      function toggleMenu(){
        const menu = document.getElementById('navMenu');
        const toggle = document.getElementById('menuToggle');
        const isOpen = menu.classList.toggle('open');
        menu.setAttribute('aria-hidden', String(!isOpen));
        toggle.setAttribute('aria-expanded', String(isOpen));
      }

      function closeMenu(){
        const menu = document.getElementById('navMenu');
        const toggle = document.getElementById('menuToggle');
        menu.classList.remove('open');
        menu.setAttribute('aria-hidden', 'true');
        toggle.setAttribute('aria-expanded', 'false');
      }

      const MONTH_TABS = {{ month_tabs|tojson }};
      const CURRENT_MONTH_KEY = "{{ current_month|default('') }}";
      let selectedMonthKey = CURRENT_MONTH_KEY;

      function monthLabelFor(key) {
        const item = MONTH_TABS.find(m => m.key === key);
        return item ? item.label : key;
      }

      function updateMonthBadges() {
        document.querySelectorAll('.month-account-row').forEach(row => {
          const badge = row.querySelector('.month-status-badge');
          const statusRaw = row.dataset.monthStatus || '{}';
          let status = {};
          try { status = JSON.parse(statusRaw); } catch (_) { status = {}; }
          const monthInfo = status[selectedMonthKey] || {};
          const category = badge?.dataset.category || 'debt';
          const isSemi = badge?.dataset.semi === 'true';
          const paid1 = monthInfo.paid1 === true;
          const paid2 = monthInfo.paid2 === true;
          let text = '';
          let bg = 'transparent';
          let color = 'inherit';
          const balance = parseFloat(badge?.dataset.balance || '0') || 0;
          if (category === 'recurring') {
            if (isSemi) {
              if (paid1 && paid2) {
                text = 'Paid';
                bg = '#16a34a';
                color = 'white';
              } else if (paid1 || paid2) {
                text = '1 of 2 paid';
                bg = '#d97706';
                color = 'white';
              }
            } else if (paid1) {
              text = 'Paid';
              bg = '#16a34a';
              color = 'white';
            }
          } else if (balance <= 0) {
            text = 'Paid off';
            bg = '#16a34a';
            color = 'white';
          } else if (paid1) {
            text = 'Paid';
            bg = '#16a34a';
            color = 'white';
          }
          badge.textContent = text;
          badge.style.background = bg;
          badge.style.color = color;
          badge.style.display = text ? 'inline-block' : 'none';
        });

        const monthLabelNodes = document.querySelectorAll('.selected-month-label');
        monthLabelNodes.forEach(node => { node.textContent = monthLabelFor(selectedMonthKey); });
      }

      (function() {
        updateMonthBadges();
      })();

      // ── Update modal ──────────────────────────────────────────────────
      const CURRENT_MONTH = "{{ current_month|default('') }}";
      const CURRENT_MONTH_LABEL = (function() {
        if (!CURRENT_MONTH) return '';
        const [y, m] = CURRENT_MONTH.split('-');
        return new Date(parseInt(y), parseInt(m) - 1, 1).toLocaleString('default', {month: 'long', year: 'numeric'});
      })();

      let currentUpdateIdx = null;
      let currentUpdateIsRecurring = false;
      let currentUpdateIsSemi = false;
      function formatIsoDate(iso) {
        if (!iso) return 'No update yet';
        const d = new Date(iso);
        if (Number.isNaN(d.getTime())) return 'No update yet';
        return d.toLocaleDateString(undefined, {month: 'long', day: 'numeric', year: 'numeric'});
      }

      function openUpdateFromRow(row) {
        if (!row) return;
        const idx = row.dataset.idx;
        const category = row.dataset.accountCategory || 'debt';
        const type = row.dataset.accountType || 'Other';
        const balance = parseFloat(row.dataset.balance || '0') || 0;
        const paid = row.dataset.paid || 'false';
        const paid2 = row.dataset.paid2 || 'false';
        const date1 = row.dataset.date1 || '';
        const date2 = row.dataset.date2 || '';
        const lastUpdated = row.dataset.lastUpdated || '';
        const minPayment = row.dataset.minPayment || '';
        const recurringAmount = row.dataset.recurringAmount || '';
        const recurringFrequency = row.dataset.recurringFrequency || '';

        currentUpdateIdx = idx;
        currentUpdateIsRecurring = category === 'recurring';
        currentUpdateIsSemi = recurringFrequency === 'semi-monthly';
        const editLink = document.getElementById('updateEditAccountLink');
        if (editLink) editLink.href = '/account/' + idx + '?return_to=settings';
        const debtFields = document.getElementById('update_debt_fields');
        const singlePaid = document.getElementById('update_paid_single');
        const semiPaid = document.getElementById('update_paid_semi');
        const title = document.getElementById('updateModalTitle');
        document.getElementById('update_account_name').textContent = row.dataset.accountName || '-';
        document.getElementById('update_account_type').textContent = type || 'Other';
        document.getElementById('update_current_balance').textContent = '$' + balance.toFixed(2);
        document.getElementById('update_last_updated').textContent = formatIsoDate(lastUpdated);
        document.getElementById('update_min_payment').innerHTML = minPayment ? '$' + parseFloat(minPayment).toFixed(2) : '—';
        if (currentUpdateIsRecurring) {
          title.textContent = 'Update recurring payment';
          debtFields.style.display = 'block';
          if (currentUpdateIsSemi) {
            singlePaid.style.display = 'none';
            semiPaid.style.display = 'flex';
            document.getElementById('update_date1_label').textContent = date1 || '?';
            document.getElementById('update_date2_label').textContent = date2 || '?';
            document.getElementById('update_mark_paid_1').checked = (paid === 'true' || row.dataset.paid1 === 'true');
            const cb2 = document.getElementById('update_mark_paid_2');
            if (cb2) cb2.checked = (paid2 === 'true');
          } else {
            singlePaid.style.display = 'block';
            semiPaid.style.display = 'none';
            document.getElementById('update_mark_paid').checked = (paid === 'true');
            document.getElementById('update_paid_month_label').textContent = monthLabelFor(selectedMonthKey);
          }
        } else {
          title.textContent = 'Update debt';
          debtFields.style.display = 'block';
          singlePaid.style.display = 'block';
          semiPaid.style.display = 'none';
          document.getElementById('update_balance').value = balance.toFixed(2);
          document.getElementById('update_type').value = type || 'Other';
          document.getElementById('update_mark_paid').checked = (paid === 'true');
          document.getElementById('update_paid_month_label').textContent = monthLabelFor(selectedMonthKey);
        }
        document.getElementById('updateResult').style.display = 'none';
        document.getElementById('updateModal').style.display = 'flex';
      }
      function closeUpdateModal(){
        document.getElementById('updateModal').style.display = 'none';
      }
      function submitUpdate(event) {
        event.preventDefault();
        if (currentUpdateIdx === null) return;
        const markPaidEl = currentUpdateIsSemi
          ? document.getElementById('update_mark_paid_1')
          : document.getElementById('update_mark_paid');
        const markPaid = markPaidEl ? markPaidEl.checked : false;
        const cb2 = document.getElementById('update_mark_paid_2');
        const markPaid2 = cb2 ? cb2.checked : false;
        const resultBox = document.getElementById('updateResult');
        const submitButton = document.getElementById('updateSubmitButton');
        submitButton.disabled = true;
        const payload = {mark_paid: markPaid};
        if (currentUpdateIsSemi) payload.mark_paid_2 = markPaid2;
        payload.target_month = selectedMonthKey;
        if (!currentUpdateIsRecurring) {
          payload.new_balance = parseFloat(document.getElementById('update_balance').value) || 0;
          payload.new_type = document.getElementById('update_type').value || 'Other';
        }
        fetch('/update/' + currentUpdateIdx, {
          method: 'POST',
          headers: {'Content-Type': 'application/json'},
          body: JSON.stringify(payload)
        }).then(async r => {
          submitButton.disabled = false;
          if (!r.ok) throw new Error('Update failed');
          const data = await r.json();
          resultBox.textContent = currentUpdateIsRecurring ? 'Payment status saved.' : `Balance updated. Current outstanding: $${parseFloat(data.balance).toFixed(2)}`;
          resultBox.style.display = 'block';
          if (!currentUpdateIsRecurring) {
            document.getElementById('update_balance').value = parseFloat(data.balance).toFixed(2);
          }
          setTimeout(() => { closeUpdateModal(); window.location.reload(); }, 1200);
        }).catch(() => {
          submitButton.disabled = false;
          resultBox.textContent = 'Save failed.';
          resultBox.style.display = 'block';
        });
      }

      // ── Monthly editor ────────────────────────────────────────────────
      function openMonthEditor(month, total){
        const section = document.getElementById('monthlyEditSection');
        if(!section) return;
        section.style.display = 'block';
        document.getElementById('edit_month_input').value = month;
        document.getElementById('edit_total_input').value = total;
        document.getElementById('edit_month_input').scrollIntoView({behavior:'smooth', block:'center'});
      }

      function hideMonthEditForm(){
        const section = document.getElementById('monthlyEditSection');
        if(section) section.style.display = 'none';
      }

      // ── Charts ────────────────────────────────────────────────────────
      function rerenderCharts(){
        renderDebtChart();
        renderTrendChart();
      }

      function renderDebtChart(){
        const chartElement = document.getElementById('debtChart');
        if(!chartElement) return;
        const chartData = [
          {% for account in accounts %}
            {name: {{ account.name|tojson }}, balance: {{ account.balance|float }}, color: {{ type_colors.get(account.type, '#64748b')|tojson }} },
          {% endfor %}
        ];
        const totalBalance = chartData.reduce((sum, item) => sum + item.balance, 0);
        const ctx = chartElement.getContext('2d');
        const canvas = ctx.canvas;
        const radius = Math.min(canvas.width, canvas.height) / 2 - 20;
        const centerX = canvas.width / 2;
        const centerY = canvas.height / 2;
        ctx.clearRect(0, 0, canvas.width, canvas.height);
        if(totalBalance <= 0){
          ctx.fillStyle = cssVar('--chart-no-data');
          ctx.font = '16px Arial';
          ctx.textAlign = 'center';
          ctx.fillText('No debt to show', centerX, centerY);
          return;
        }

        let startAngle = -Math.PI / 2;
        const legend = document.getElementById('chartLegend');
        legend.innerHTML = '';
        chartData.filter(item => item.balance > 0).forEach(item => {
          const sliceAngle = (item.balance / totalBalance) * Math.PI * 2;
          ctx.beginPath();
          ctx.moveTo(centerX, centerY);
          ctx.arc(centerX, centerY, radius, startAngle, startAngle + sliceAngle);
          ctx.closePath();
          ctx.fillStyle = item.color;
          ctx.fill();
          startAngle += sliceAngle;

          const percent = ((item.balance / totalBalance) * 100).toFixed(1);
          const legendItem = document.createElement('div');
          legendItem.className = 'chart-legend-item';
          legendItem.innerHTML = `<span class="chart-legend-color" style="background:${item.color}"></span><strong>${item.name}</strong> ${percent}%`;
          legend.appendChild(legendItem);
        });

        ctx.fillStyle = cssVar('--chart-label');
        ctx.font = 'bold 18px Arial';
        ctx.textAlign = 'center';
        ctx.fillText('Debt %', centerX, centerY - 10);
        ctx.font = 'bold 16px Arial';
        ctx.fillText(`${totalBalance.toFixed(2)}`, centerX, centerY + 18);
      }

      function renderTrendChart(){
        const chartElement = document.getElementById('trendChart');
        if(!chartElement) return;
        const trendData = {{ trend_data|default([])|tojson }};
        const ctx = chartElement.getContext('2d');
        const width = chartElement.width;
        const height = chartElement.height;
        ctx.clearRect(0, 0, width, height);
        if(!trendData.length){
          ctx.fillStyle = cssVar('--chart-no-data');
          ctx.font = '16px Arial';
          ctx.textAlign = 'center';
          ctx.fillText('No trend data yet', width / 2, height / 2);
          return;
        }

        const labels = trendData.map(item => item.month);
        const values = trendData.map(item => item.total);
        const maxValue = Math.max(...values);
        const minValue = Math.min(...values);
        const padding = 40;
        const chartWidth = width - padding * 2;
        const chartHeight = height - padding * 2;
        const range = maxValue - minValue || 1;

        ctx.strokeStyle = cssVar('--chart-grid');
        ctx.lineWidth = 1;
        for(let i = 0; i <= 4; i++){
          const y = padding + (chartHeight / 4) * i;
          ctx.beginPath();
          ctx.moveTo(padding, y);
          ctx.lineTo(width - padding, y);
          ctx.stroke();
        }

        const points = values.map((value, index) => {
          const x = padding + (chartWidth * index) / (values.length - 1 || 1);
          const y = padding + chartHeight * (1 - (value - minValue) / range);
          return {x, y, value};
        });

        ctx.beginPath();
        ctx.strokeStyle = '#2563eb';
        ctx.lineWidth = 3;
        points.forEach((point, index) => {
          if(index === 0) ctx.moveTo(point.x, point.y);
          else ctx.lineTo(point.x, point.y);
        });
        ctx.stroke();

        ctx.fillStyle = 'rgba(59,130,246,0.15)';
        ctx.beginPath();
        points.forEach((point, index) => {
          if(index === 0) ctx.moveTo(point.x, point.y);
          else ctx.lineTo(point.x, point.y);
        });
        ctx.lineTo(points[points.length - 1].x, height - padding);
        ctx.lineTo(points[0].x, height - padding);
        ctx.closePath();
        ctx.fill();

        ctx.fillStyle = '#2563eb';
        points.forEach(point => {
          ctx.beginPath();
          ctx.arc(point.x, point.y, 4, 0, Math.PI * 2);
          ctx.fill();
        });

        ctx.fillStyle = cssVar('--chart-axis');
        ctx.font = '12px Arial';
        ctx.textAlign = 'center';
        labels.forEach((label, index) => {
          const point = points[index];
          ctx.fillText(label, point.x, height - padding + 16);
        });

        ctx.fillStyle = cssVar('--chart-label');
        ctx.font = 'bold 14px Arial';
        ctx.textAlign = 'left';
        ctx.fillText(`From ${labels[0]} to ${labels[labels.length - 1]}`, padding, padding - 10);
      }

      // ── Settings swipe ────────────────────────────────────────────────
      function archiveAccount(idx) {
        fetch('/archive/' + idx, {method:'POST'}).then(r => { if(r.ok) window.location.reload(); });
      }
      function unarchiveAccount(idx) {
        fetch('/unarchive/' + idx, {method:'POST'}).then(r => { if(r.ok) window.location.reload(); });
      }

      function enableSettingsListSwipe(){
        document.querySelectorAll('.settings-account-row').forEach(row => {
          const deleteButton = row.querySelector('.settings-account-delete');
          let startX = 0;

          row.addEventListener('touchstart', function(event){
            startX = event.touches[0].clientX;
          });
          row.addEventListener('touchmove', function(event){
            const currentX = event.touches[0].clientX;
            const deltaX = currentX - startX;
            if(deltaX < -20){
              row.classList.add('swiped');
            } else if(deltaX > 20) {
              row.classList.remove('swiped');
            }
          });

          deleteButton.addEventListener('click', function(){
            if(!confirm('Delete this account? This action cannot be undone.')) return;
            fetch('/delete/' + row.dataset.idx, { method: 'POST' }).then(r => {
              if(r.ok) window.location.reload();
              else alert('Delete failed');
            }).catch(() => alert('Delete failed'));
          });
        });
      }

      function toggleAddFrequency(freq) {
        const isSemi = freq === 'semi-monthly';
        const single = document.getElementById('add_due_date_single');
        const semi = document.getElementById('add_due_date_semi');
        if (single) single.style.display = isSemi ? 'none' : 'block';
        if (semi) semi.style.display = isSemi ? 'block' : 'none';
      }

      function toggleDetailFrequency(freq) {
        const isSemi = freq === 'semi-monthly';
        const single = document.getElementById('detail_due_single');
        const semi = document.getElementById('detail_due_semi');
        if (single) single.style.display = isSemi ? 'none' : 'block';
        if (semi) semi.style.display = isSemi ? 'block' : 'none';
      }

      function toggleAddCategory(cat) {
        const isRecurring = cat === 'recurring';
        document.getElementById('add_debt_fields').style.display = isRecurring ? 'none' : 'block';
        document.getElementById('add_recurring_fields').style.display = isRecurring ? 'flex' : 'none';
        document.getElementById('add_interest_field').style.display = isRecurring ? 'none' : 'block';
        document.getElementById('add_type_debt').style.display = isRecurring ? 'none' : 'block';
        document.getElementById('add_type_debt').disabled = isRecurring;
        document.getElementById('add_type_recurring').style.display = isRecurring ? 'block' : 'none';
        document.getElementById('add_type_recurring').disabled = !isRecurring;
        if (!isRecurring) toggleAddFrequency('monthly');
      }

      function toggleCustomDueDate(selectEl, customEl){
        if(!selectEl || !customEl) return;
        customEl.style.display = selectEl.value === 'custom' ? 'block' : 'none';
      }
      function wireDueDateSelect(selectId, customId){
        const selectEl = document.getElementById(selectId);
        const customEl = document.getElementById(customId);
        if(!selectEl || !customEl) return;
        selectEl.addEventListener('change', function(){ toggleCustomDueDate(selectEl, customEl); });
        toggleCustomDueDate(selectEl, customEl);
      }
      document.addEventListener('DOMContentLoaded', function(){
        renderDebtChart();
        renderTrendChart();
        enableSettingsListSwipe();
        wireDueDateSelect('detail_due_date_select', 'detail_due_date_custom');
        document.querySelectorAll('.month-account-row').forEach(row => {
          const openRow = () => openUpdateFromRow(row);
          row.addEventListener('click', function(event) {
            if (event.target.closest('a, button, input, select, textarea')) return;
            openRow();
          });
          row.addEventListener('keydown', function(event) {
            if (event.key === 'Enter' || event.key === ' ') {
              event.preventDefault();
              openRow();
            }
          });
        });
      });
      document.addEventListener('keydown', function(e){ if(e.key === 'Escape'){ closeUpdateModal(); } });
    </script>
  </body>
</html>
"""


def load_data():
    init_storage()
    conn = get_db_connection()
    try:
        row = conn.execute("SELECT payload FROM app_state WHERE id = 1").fetchone()
        if row and row["payload"]:
            try:
                return json.loads(row["payload"])
            except (ValueError, json.JSONDecodeError):
                return {"accounts": []}
        return {"accounts": []}
    finally:
        conn.close()


def save_data(data):
    init_storage()
    payload = json.dumps(data, indent=2)
    conn = get_db_connection()
    try:
        conn.execute("UPDATE app_state SET payload = ? WHERE id = 1", (payload,))
        conn.commit()
    finally:
        conn.close()


def parse_timestamp(value):
    if not value:
        return None
    value = value.strip()
    if not value:
        return None
    try:
        return datetime.fromisoformat(value)
    except ValueError:
        try:
            return datetime.strptime(value, '%Y-%m-%d')
        except ValueError:
            try:
                return datetime.strptime(value, '%Y-%m')
            except ValueError:
                return None


def make_history_signature(entry):
    return '|'.join([
        str(entry.get('timestamp', '')),
        str(entry.get('action', '')),
        str(entry.get('amount', '')),
        str(entry.get('prev_amount', '')),
        str(entry.get('payment', '')),
        str(entry.get('new_type', '')),
        str(entry.get('new_due_date', '')),
        str(entry.get('new_interest_rate', '')),
        str(entry.get('old_name', '')),
        str(entry.get('new_name', ''))
    ])


def compute_next_due(due_date_str, today):
    """Return (next_due date, days_until int) or (None, None) if unknown/empty."""
    if not due_date_str or due_date_str.strip().lower() in ('unknown', ''):
        return None, None
    d = due_date_str.strip().lower()
    if d == '1st':
        day = 1
    elif d == '15th':
        day = 15
    elif d == 'last':
        day = None
    else:
        try:
            day = int(d)
        except ValueError:
            return None, None

    def date_for(y, m):
        if day is None:
            return date(y, m, monthrange(y, m)[1])
        return date(y, m, min(day, monthrange(y, m)[1]))

    candidate = date_for(today.year, today.month)
    if candidate < today:
        m, y = today.month % 12 + 1, today.year + (1 if today.month == 12 else 0)
        candidate = date_for(y, m)

    return candidate, (candidate - today).days


def build_history_rows(data):
    rows = []
    for account in data.get('accounts', []):
        for entry in account.get('history', []):
            rows.append({
                'account_name': account.get('name', ''),
                'timestamp': entry.get('timestamp', ''),
                'action': entry.get('action', ''),
                'amount': entry.get('amount', ''),
                'prev_amount': entry.get('prev_amount', ''),
                'payment': entry.get('payment', ''),
                'old_type': entry.get('old_type', ''),
                'new_type': entry.get('new_type', ''),
                'old_name': entry.get('old_name', ''),
                'new_name': entry.get('new_name', ''),
                'old_due_date': entry.get('old_due_date', ''),
                'new_due_date': entry.get('new_due_date', ''),
                'old_interest_rate': entry.get('old_interest_rate', ''),
                'new_interest_rate': entry.get('new_interest_rate', ''),
                'account_type': account.get('type', ''),
                'due_date': account.get('due_date', ''),
                'interest_rate': account.get('interest_rate', '')
            })
    return rows


@app.route("/", methods=["GET"])
async def index(request: Request):
  data = load_data()
  raw_accounts = data.get("accounts", [])
  settings = data.get("settings", {})

  # allow sorting via query params; default to days until due ascending
  sort_by = request.query_params.get('sort', 'due_date')
  order = request.query_params.get('order', 'asc')
  hide_paid = request.query_params.get('hide_paid', '0') == '1'

  show_zero = settings.get('show_zero', False)

  today = datetime.now().date()
  current_month = today.strftime('%Y-%m')
  def month_shift(month_key, delta):
    current = datetime.strptime(month_key + '-01', '%Y-%m-%d')
    year = current.year + ((current.month - 1 + delta) // 12)
    month = ((current.month - 1 + delta) % 12) + 1
    return datetime(year, month, 1).strftime('%Y-%m')

  previous_month = month_shift(current_month, -1)
  next_month = month_shift(current_month, 1)

  # Build a view list that preserves original indices so actions map correctly
  indexed = [ (i, a) for i, a in enumerate(raw_accounts) if not a.get('archived') ]

  if not show_zero:
    indexed = [ (i, a) for (i, a) in indexed if a.get('category') == 'recurring' or float(a.get('balance', 0) or 0) != 0]

  if hide_paid:
    def is_fully_paid(a):
      if a.get('recurring_frequency') == 'semi-monthly':
        return a.get('paid_month') == current_month and a.get('paid_month_2') == current_month
      return a.get('paid_month') == current_month
    indexed = [ (i, a) for (i, a) in indexed if not is_fully_paid(a) ]

  def nearest_due(acct):
    nd1, du1 = compute_next_due(acct.get('due_date', ''), today)
    if acct.get('due_date_2') and acct.get('recurring_frequency') == 'semi-monthly':
      nd2, du2 = compute_next_due(acct.get('due_date_2', ''), today)
      if nd2 is not None and (nd1 is None or du2 < du1):
        return nd2, du2
    return nd1, du1

  def days_until_key(t):
    _, days = nearest_due(t[1])
    return days if days is not None else 9999

  def payment_priority(acct):
    return 0 if not acct.get('paid_this_month') else 1

  def last_updated_key(acct):
    ts = parse_timestamp(acct.get('last_updated'))
    return ts or datetime.min

  if sort_by == 'amount':
    indexed = sorted(indexed, key=lambda t: float(t[1].get('balance', 0) or 0), reverse=(order == 'desc'))
  elif sort_by == 'name':
    indexed = sorted(indexed, key=lambda t: (t[1].get('name') or '').lower(), reverse=(order == 'desc'))
  elif sort_by == 'type':
    indexed = sorted(indexed, key=lambda t: (t[1].get('type') or '').lower(), reverse=(order == 'desc'))
  elif sort_by == 'due_date':
    indexed = sorted(
      indexed,
      key=lambda t: (payment_priority(t[1]), days_until_key(t), last_updated_key(t[1])),
      reverse=False
    )

  view_mode = request.query_params.get('view', 'list')
  if view_mode not in ('list', 'chart'):
    view_mode = 'list'

  current_month_label = today.strftime('%B %Y')
  previous_month_label = datetime.strptime(previous_month + '-01', '%Y-%m-%d').strftime('%B %Y')
  next_month_label = datetime.strptime(next_month + '-01', '%Y-%m-%d').strftime('%B %Y')
  month_tabs = [
    {'key': previous_month, 'label': previous_month_label},
    {'key': current_month, 'label': current_month_label},
    {'key': next_month, 'label': next_month_label},
  ]

  accounts = []
  for orig_idx, acct in indexed:
    acct_copy = dict(acct)
    acct_copy['orig_idx'] = orig_idx
    next_date, days_until = nearest_due(acct)
    acct_copy['days_until'] = days_until
    acct_copy['next_due_date'] = (next_date.strftime('%b ') + str(next_date.day)) if next_date else None
    if acct.get('due_date_2') and acct.get('recurring_frequency') == 'semi-monthly':
      _, du2 = compute_next_due(acct.get('due_date_2', ''), today)
      acct_copy['due_dates_label'] = (acct.get('due_date', '') + ' & ' + acct.get('due_date_2', '')) if du2 is not None else None
    else:
      acct_copy['due_dates_label'] = None
    is_semi = acct.get('recurring_frequency') == 'semi-monthly'
    paid1 = acct.get('paid_month') == current_month
    paid2 = acct.get('paid_month_2') == current_month
    acct_copy['paid_1_this_month'] = paid1
    acct_copy['paid_2_this_month'] = paid2
    acct_copy['paid_this_month'] = (paid1 and paid2) if is_semi else paid1
    acct_copy['last_updated'] = acct.get('last_updated')
    acct_copy['paid_status_by_month'] = {
      previous_month: {
        'paid1': acct.get('paid_month') == previous_month,
        'paid2': acct.get('paid_month_2') == previous_month,
      },
      current_month: {
        'paid1': acct.get('paid_month') == current_month,
        'paid2': acct.get('paid_month_2') == current_month,
      },
      next_month: {
        'paid1': acct.get('paid_month') == next_month,
        'paid2': acct.get('paid_month_2') == next_month,
      },
    }
    accounts.append(acct_copy)

  total = sum(a.get("balance", 0) for a in accounts if a.get('category', 'debt') == 'debt')

  remaining_this_month = 0
  for a in raw_accounts:
    if a.get('archived'):
      continue
    if a.get('category') == 'recurring':
      freq = a.get('recurring_frequency', 'monthly')
      amt = float(a.get('recurring_amount', 0) or 0)
      if freq == 'semi-monthly':
        if a.get('paid_month') != current_month:
          remaining_this_month += amt
        if a.get('paid_month_2') != current_month:
          remaining_this_month += amt
      else:
        if a.get('paid_month') != current_month:
          remaining_this_month += amt
    elif a.get('category', 'debt') == 'debt':
      mp = float(a.get('min_payment', 0) or 0)
      if mp > 0 and a.get('paid_month') != current_month:
        remaining_this_month += mp

  def to_monthly(amount, freq):
    if freq == 'yearly': return amount / 12
    if freq == 'quarterly': return amount / 3
    if freq == 'weekly': return amount * 52 / 12
    if freq == 'semi-monthly': return amount * 2
    return amount

  monthly_recurring = sum(
    to_monthly(float(a.get('recurring_amount', 0) or 0), a.get('recurring_frequency', 'monthly'))
    for a in raw_accounts if a.get('category') == 'recurring' and not a.get('archived')
  )

  type_colors = {
      'Credit Card': '#ef4444',
      'Car Loan': '#10b981',
      'Mortgage': '#2563eb',
      'Student Loan': '#f59e0b',
      'Personal Loan': '#8b5cf6',
      'Subscription': '#06b6d4',
      'Streaming': '#f43f5e',
      'Utility': '#84cc16',
      'Insurance': '#fb923c',
      'Other': '#64748b'
  }

  timeline = []
  for account in raw_accounts:
    for entry in account.get('history', []):
      ts = entry.get('timestamp')
      if not ts:
        continue
      try:
        dt = datetime.fromisoformat(ts)
      except ValueError:
        continue
      timeline.append({
        'timestamp': dt,
        'account_name': entry.get('account_name', account.get('name', 'Unknown')),
        'balance': float(entry.get('amount', 0) or 0)
      })
  timeline.sort(key=lambda e: e['timestamp'])

  trend_data = []
  if timeline:
    monthly_totals = {}
    current_balances = {}
    for event in timeline:
      current_balances[event['account_name']] = event['balance']
      month_key = event['timestamp'].strftime('%Y-%m')
      monthly_totals[month_key] = sum(current_balances.values())

    first_month = min(monthly_totals.keys())
    last_month = max(monthly_totals.keys())
    def month_iter(start, end):
      current = datetime.fromisoformat(start + '-01')
      end_dt = datetime.fromisoformat(end + '-01')
      while current <= end_dt:
        yield current.strftime('%Y-%m')
        year = current.year + (current.month // 12)
        month = current.month % 12 + 1
        current = current.replace(year=year, month=month)

    last_total = 0
    for month_key in month_iter(first_month, last_month):
      last_total = monthly_totals.get(month_key, last_total)
      trend_data.append({'month': month_key, 'total': last_total})

  return render_template_string(TEMPLATE, page='tasks', accounts=accounts, total=total, monthly_recurring=monthly_recurring, remaining_this_month=remaining_this_month, sort_by=sort_by, order=order, view_mode=view_mode, type_colors=type_colors, show_zero=show_zero, trend_data=trend_data, current_month=current_month, current_month_label=current_month_label, previous_month=previous_month, next_month=next_month, month_tabs=month_tabs, hide_paid=hide_paid)


@app.route('/dashboard', methods=['GET'])
async def dashboard_page(request: Request):
  data = load_data()
  raw_accounts = data.get("accounts", [])
  today = datetime.now().date()
  current_month = today.strftime('%Y-%m')
  current_month_label = today.strftime('%B %Y')

  accounts = [dict(a, orig_idx=i) for i, a in enumerate(raw_accounts) if not a.get('archived')]
  total = sum(a.get("balance", 0) for a in accounts if a.get('category', 'debt') == 'debt')
  remaining_this_month = 0
  for a in raw_accounts:
    if a.get('archived'):
      continue
    if a.get('category') == 'recurring':
      freq = a.get('recurring_frequency', 'monthly')
      amt = float(a.get('recurring_amount', 0) or 0)
      if freq == 'semi-monthly':
        if a.get('paid_month') != current_month:
          remaining_this_month += amt
        if a.get('paid_month_2') != current_month:
          remaining_this_month += amt
      else:
        if a.get('paid_month') != current_month:
          remaining_this_month += amt
    elif a.get('category', 'debt') == 'debt':
      mp = float(a.get('min_payment', 0) or 0)
      if mp > 0 and a.get('paid_month') != current_month:
        remaining_this_month += mp

  monthly_recurring = sum(
    (float(a.get('recurring_amount', 0) or 0) if a.get('category') == 'recurring' else 0)
    for a in raw_accounts if not a.get('archived')
  )

  type_colors = {
      'Credit Card': '#ef4444',
      'Car Loan': '#10b981',
      'Mortgage': '#2563eb',
      'Student Loan': '#f59e0b',
      'Personal Loan': '#8b5cf6',
      'Subscription': '#06b6d4',
      'Streaming': '#f43f5e',
      'Utility': '#84cc16',
      'Insurance': '#fb923c',
      'Other': '#64748b'
  }

  timeline = []
  for account in raw_accounts:
    for entry in account.get('history', []):
      ts = entry.get('timestamp')
      if not ts:
        continue
      try:
        dt = datetime.fromisoformat(ts)
      except ValueError:
        continue
      timeline.append({
        'timestamp': dt,
        'account_name': entry.get('account_name', account.get('name', 'Unknown')),
        'balance': float(entry.get('amount', 0) or 0)
      })
  timeline.sort(key=lambda e: e['timestamp'])

  trend_data = []
  if timeline:
    monthly_totals = {}
    current_balances = {}
    for entry in timeline:
      month_key = entry['timestamp'].strftime('%Y-%m')
      current_balances[entry['account_name']] = entry['balance']
      monthly_totals[month_key] = sum(current_balances.values())
    first_month = min(monthly_totals.keys())
    last_month = max(monthly_totals.keys())
    start = datetime.strptime(first_month + '-01', '%Y-%m-%d')
    end = datetime.strptime(last_month + '-01', '%Y-%m-%d')
    month = start
    last_total = 0
    while month <= end:
      month_key = month.strftime('%Y-%m')
      if month_key in monthly_totals:
        last_total = monthly_totals[month_key]
      trend_data.append({'month': month_key, 'total': last_total})
      month = datetime(month.year + (month.month // 12), (month.month % 12) + 1, 1)

  return render_template_string(TEMPLATE, page='dashboard', accounts=accounts, total=total, monthly_recurring=monthly_recurring, remaining_this_month=remaining_this_month, type_colors=type_colors, trend_data=trend_data, current_month=current_month, current_month_label=current_month_label)


@app.route('/settings', methods=['GET', 'POST'])
async def settings_page(request: Request):
  data = load_data()
  raw_accounts = data.get('accounts', [])
  settings = data.setdefault('settings', {})

  if request.method == 'POST':
    form = await read_form_data(request)
    action = form.get('action')
    if action == 'update_settings':
      settings['show_zero'] = form.get('show_zero') == '1'
      data['settings'] = settings
      save_data(data)
      return redirect(url_for('settings_page'))
    if action == 'add_account':
      new_name = form.get('new_name', '').strip()
      new_category = form.get('new_category', 'debt')
      if new_name:
        new_type = form.get('new_type', 'Other')
        new_owner = form.get('new_owner', '').strip()
        freq = form.get('new_recurring_frequency', 'monthly')
        if new_category == 'recurring' and freq == 'semi-monthly':
          new_due_date = form.get('new_semi_date_1', '').strip()
          new_due_date_2 = form.get('new_semi_date_2', '').strip()
        else:
          new_due_date_2 = None
          due_date_type = form.get('new_due_date_type')
          if due_date_type == 'custom':
            new_due_date = form.get('new_due_date', '').strip()
          elif due_date_type in ('1st', '15th', 'last', 'unknown'):
            new_due_date = due_date_type
          else:
            new_due_date = form.get('new_due_date', '').strip()
        new_acct = {
          'name': new_name,
          'category': new_category,
          'type': new_type,
          'owner': new_owner,
          'due_date': new_due_date,
          'due_date_2': new_due_date_2,
          'history': [{
            'action': 'Created account',
            'account_name': new_name,
            'timestamp': datetime.utcnow().isoformat()
          }]
        }
        if new_category == 'recurring':
          try:
            new_acct['recurring_amount'] = round(float(form.get('new_recurring_amount', 0) or 0), 2)
          except ValueError:
            new_acct['recurring_amount'] = 0.0
          new_acct['recurring_frequency'] = form.get('new_recurring_frequency', 'monthly')
          new_acct['balance'] = 0.0
          new_acct['interest_rate'] = None
        else:
          try:
            new_acct['balance'] = round(float(form.get('new_balance', '0') or 0), 2)
          except ValueError:
            new_acct['balance'] = 0.0
          try:
            ir = float(form.get('new_interest_rate', '') or 0)
            new_acct['interest_rate'] = round(ir, 2) if ir != 0 else None
          except ValueError:
            new_acct['interest_rate'] = None
        raw_accounts.append(new_acct)
        data['accounts'] = raw_accounts
        save_data(data)
      return redirect(url_for('settings_page'))

  show_zero = settings.get('show_zero', False)
  status = request.query_params.get('status', '')
  sort_by = request.query_params.get('sort', 'amount')
  order = request.query_params.get('order', 'desc')
  total = sum(account.get('balance', 0) for account in raw_accounts)
  type_colors = {
      'Credit Card': '#ef4444',
      'Car Loan': '#10b981',
      'Mortgage': '#2563eb',
      'Student Loan': '#f59e0b',
      'Personal Loan': '#8b5cf6',
      'Other': '#64748b'
  }
  all_with_idx = [dict(a, orig_idx=i) for i, a in enumerate(raw_accounts)]
  settings_accounts = sorted([a for a in all_with_idx if not a.get('archived')], key=lambda a: (a.get('name') or '').lower())
  archived_accounts = sorted([a for a in all_with_idx if a.get('archived')], key=lambda a: (a.get('name') or '').lower())
  return render_template_string(
    TEMPLATE,
    page='settings',
    accounts=settings_accounts,
    archived_accounts=archived_accounts,
    total=total,
    sort_by=sort_by,
    order=order,
    type_colors=type_colors,
    show_zero=show_zero,
    status=status,
    month_tabs=[],
    trend_data=[],
    current_month='',
    current_month_label='',
  )


@app.route('/account/{idx}', methods=['GET'])
async def account_page(idx: int, request: Request):
  data = load_data()
  accounts = data.get('accounts', [])
  if idx < 0 or idx >= len(accounts):
    abort(404)
  account = accounts[idx]
  history = sorted(account.get('history', []), key=lambda e: parse_timestamp(e.get('timestamp')) or datetime.min)
  monthly_totals = {}
  for entry in history:
    ts = parse_timestamp(entry.get('timestamp'))
    if not ts:
      continue
    month_key = ts.strftime('%Y-%m')
    monthly_totals[month_key] = float(entry.get('amount', 0) or 0)

  monthly_list = []
  if monthly_totals:
    first_month = min(monthly_totals.keys())
    last_month = max(monthly_totals.keys())
    current = datetime.fromisoformat(first_month + '-01')
    end = datetime.fromisoformat(last_month + '-01')
    last_value = 0.0
    while current <= end:
      month_key = current.strftime('%Y-%m')
      if month_key in monthly_totals:
        last_value = monthly_totals[month_key]
      monthly_list.append({'month': month_key, 'total': last_value})
      year = current.year + (current.month // 12)
      month = current.month % 12 + 1
      current = current.replace(year=year, month=month)

  return_to = request.query_params.get('return_to', 'settings')
  return render_template_string(
    TEMPLATE,
    page='account_detail',
    account=account,
    monthly_totals=monthly_list,
    idx=idx,
    return_to=return_to,
    month_tabs=[],
    trend_data=[],
    current_month='',
    current_month_label='',
    accounts=[],
    archived_accounts=[],
    total=0,
    sort_by='amount',
    order='desc',
    type_colors={},
    show_zero=False,
    status='',
  )


@app.route('/account/{idx}/update_month', methods=['POST'])
async def update_month(idx: int, request: Request):
  data = load_data()
  accounts = data.get('accounts', [])
  if idx < 0 or idx >= len(accounts):
    abort(404)
  account = accounts[idx]
  form = await read_form_data(request)
  raw_text = form.get('monthly_totals', '').strip()
  parsed_entries = []
  if raw_text:
    lines = [line.strip() for line in raw_text.splitlines() if line.strip()]
    for line in lines:
      parts = [part for part in re.split(r'[\s,]+', line) if part]
      if len(parts) < 2:
        continue
      month_value = parts[0]
      total_value = parts[1]
      timestamp = parse_timestamp(month_value)
      if timestamp is None:
        continue
      try:
        total_amount = float(total_value)
      except ValueError:
        continue
      parsed_entries.append((timestamp, round(total_amount, 2)))
  else:
    month_value = form.get('month', '').strip()
    total_value = form.get('total', '').strip()
    if month_value and total_value:
      timestamp = parse_timestamp(month_value)
      if timestamp is not None:
        try:
          total_amount = float(total_value)
        except ValueError:
          total_amount = None
        if total_amount is not None:
          parsed_entries.append((timestamp, round(total_amount, 2)))

  if not parsed_entries:
    return redirect(url_for('account_page', idx=idx))

  parsed_entries.sort(key=lambda item: item[0])
  latest_existing = max(
    (parse_timestamp(e.get('timestamp')) or datetime.min for e in account.get('history', [])),
    default=datetime.min
  )

  for timestamp, total_amount in parsed_entries:
    existing_entries = [e for e in account.get('history', []) if parse_timestamp(e.get('timestamp')) and parse_timestamp(e.get('timestamp')).strftime('%Y-%m') == timestamp.strftime('%Y-%m')]
    if existing_entries:
      target = max(existing_entries, key=lambda e: parse_timestamp(e.get('timestamp')))
      target['amount'] = round(total_amount, 2)
      target['action'] = f'Updated total for {timestamp.strftime("%Y-%m")}'
      target['payment'] = None
      target['old_type'] = account.get('type', 'Other')
      target['new_type'] = account.get('type', 'Other')
      target['timestamp'] = timestamp.isoformat()
    else:
      account.setdefault('history', []).append({
        'action': f'Updated total for {timestamp.strftime("%Y-%m")}',
        'account_name': account.get('name', 'Unknown'),
        'amount': round(total_amount, 2),
        'prev_amount': None,
        'payment': None,
        'old_type': account.get('type', 'Other'),
        'new_type': account.get('type', 'Other'),
        'timestamp': timestamp.isoformat()
      })

    if timestamp >= latest_existing:
      account['balance'] = round(total_amount, 2)
      latest_existing = timestamp

  account['history'] = sorted(
    account.get('history', []),
    key=lambda e: parse_timestamp(e.get('timestamp')) or datetime.min
  )
  data['accounts'] = accounts
  save_data(data)
  return redirect(url_for('account_page', idx=idx))


@app.route('/export', methods=['GET'])
def export_data():
  data = load_data()
  today_str = datetime.now().strftime('%Y-%m-%d')
  payload = json.dumps(data, indent=2)
  response = Response(payload, mimetype='application/json')
  response.headers['Content-Disposition'] = f'attachment; filename=debt-tracker-{today_str}.json'
  return response


@app.route('/export_excel', methods=['GET'])
def export_excel():
  data = load_data()
  today_str = datetime.now().strftime('%Y-%m-%d')
  workbook = Workbook()
  workbook.remove(workbook.active)

  accounts = data.get('accounts', [])
  settings = data.get('settings', {})
  history_rows = build_history_rows(data)
  account_rows = []
  archived_rows = []
  monthly_rows = []

  category_totals = {}
  account_type_totals = {}
  monthly_history = {}
  for idx, account in enumerate(accounts):
    balance = float(account.get('balance', 0) or 0)
    recurring_amount = float(account.get('recurring_amount', 0) or 0)
    min_payment = float(account.get('min_payment', 0) or 0)
    value_for_chart = recurring_amount if account.get('category') == 'recurring' else balance
    account_type = account.get('type') or 'Other'
    category = account.get('category') or 'debt'
    category_totals[category] = category_totals.get(category, 0) + value_for_chart
    account_type_totals[account_type] = account_type_totals.get(account_type, 0) + value_for_chart

    flat_account = {k: excel_safe(v) for k, v in account.items() if k != 'history'}
    flat_account['index'] = idx
    flat_account['history_count'] = len(account.get('history', []))
    flat_account['monthly_value'] = value_for_chart
    flat_account['effective_payment'] = recurring_amount if category == 'recurring' else min_payment
    flat_account['is_archived'] = bool(account.get('archived'))
    account_rows.append(flat_account)
    if account.get('archived'):
      archived_rows.append(flat_account)

    for entry in account.get('monthly_totals', []):
      month_key = entry.get('month', '')
      amount = float(entry.get('total', 0) or 0)
      monthly_rows.append({
        'account_index': idx,
        'account_name': account.get('name', ''),
        'month': month_key,
        'total': amount,
        'type': account_type,
        'due_date': account.get('due_date', ''),
        'interest_rate': account.get('interest_rate', ''),
      })
      monthly_history[month_key] = monthly_history.get(month_key, 0) + amount

  summary_sheet = workbook.create_sheet(title='Dashboard')
  summary_sheet['A1'] = 'Debt Tracker Summary'
  summary_sheet['A1'].font = Font(bold=True, size=18, color='1F2937')
  summary_sheet['A3'] = 'Total accounts'
  summary_sheet['B3'] = len(accounts)
  summary_sheet['A4'] = 'Active accounts'
  summary_sheet['B4'] = len([a for a in accounts if not a.get('archived')])
  summary_sheet['A5'] = 'Archived accounts'
  summary_sheet['B5'] = len([a for a in accounts if a.get('archived')])
  summary_sheet['A6'] = 'Total outstanding'
  summary_sheet['B6'] = sum(float(a.get('balance', 0) or 0) for a in accounts if a.get('category') != 'recurring' and not a.get('archived'))
  summary_sheet['A7'] = 'Monthly recurring value'
  summary_sheet['B7'] = sum(float(a.get('recurring_amount', 0) or 0) for a in accounts if a.get('category') == 'recurring' and not a.get('archived'))
  summary_sheet['A8'] = 'Monthly minimum payments'
  summary_sheet['B8'] = sum(float(a.get('min_payment', 0) or 0) for a in accounts if a.get('category') != 'recurring' and not a.get('archived'))
  summary_sheet['D3'] = 'Category'
  summary_sheet['E3'] = 'Total'
  row = 4
  for label, value in sorted(category_totals.items()):
    summary_sheet[f'D{row}'] = label.title()
    summary_sheet[f'E{row}'] = value
    row += 1
  summary_sheet['D10'] = 'Type'
  summary_sheet['E10'] = 'Total'
  row = 11
  for label, value in sorted(account_type_totals.items(), key=lambda item: item[1], reverse=True):
    summary_sheet[f'D{row}'] = label
    summary_sheet[f'E{row}'] = value
    row += 1
  summary_sheet['G3'] = 'Month'
  summary_sheet['H3'] = 'Total'
  row = 4
  for month_key in sorted(monthly_history.keys()):
    summary_sheet[f'G{row}'] = month_key
    summary_sheet[f'H{row}'] = monthly_history[month_key]
    row += 1

  for cell_range in ['A3:B8', 'D3:E20', 'G3:H20']:
    for row_cells in summary_sheet[cell_range]:
      for cell in row_cells:
        cell.border = Border(bottom=Side(style='thin', color='D1D5DB'))

  for cell in summary_sheet[3]:
    if cell.value:
      cell.font = Font(bold=True, color='FFFFFF')
      cell.fill = PatternFill('solid', fgColor='2563EB')
  for cell in summary_sheet[10]:
    if cell.value:
      cell.font = Font(bold=True, color='FFFFFF')
      cell.fill = PatternFill('solid', fgColor='2563EB')

  pie = PieChart()
  labels = Reference(summary_sheet, min_col=4, min_row=4, max_row=3 + len(category_totals))
  data_ref = Reference(summary_sheet, min_col=5, min_row=3, max_row=3 + len(category_totals))
  pie.add_data(data_ref, titles_from_data=True)
  pie.set_categories(labels)
  pie.title = 'Debt by Category'
  pie.height = 7
  pie.width = 9
  pie.dataLabels = DataLabelList()
  pie.dataLabels.showPercent = True
  summary_sheet.add_chart(pie, 'J3')

  bar = BarChart()
  bar.type = 'bar'
  bar.style = 10
  bar.title = 'Accounts by Type'
  bar.y_axis.title = 'Type'
  bar.x_axis.title = 'Total'
  bar.height = 7
  bar.width = 11
  data_ref = Reference(summary_sheet, min_col=5, min_row=10, max_row=10 + len(account_type_totals))
  cats = Reference(summary_sheet, min_col=4, min_row=11, max_row=10 + len(account_type_totals))
  bar.add_data(data_ref, titles_from_data=True)
  bar.set_categories(cats)
  summary_sheet.add_chart(bar, 'J20')

  line = LineChart()
  line.title = 'Monthly Totals'
  line.y_axis.title = 'Amount'
  line.x_axis.title = 'Month'
  line.height = 7
  line.width = 13
  if monthly_history:
    line_data = Reference(summary_sheet, min_col=8, min_row=3, max_row=3 + len(monthly_history))
    line_cats = Reference(summary_sheet, min_col=7, min_row=4, max_row=3 + len(monthly_history))
    line.add_data(line_data, titles_from_data=True)
    line.set_categories(line_cats)
    summary_sheet.add_chart(line, 'J37')

  for sheet_name, rows in [
    ('Accounts', account_rows),
    ('Archived', archived_rows),
    ('Monthly Totals', monthly_rows),
    ('History', history_rows),
    ('Settings', [{'key': k, 'value': excel_safe(v)} for k, v in settings.items()]),
  ]:
    sheet = sheet_from_rows(workbook, sheet_name, rows)
    style_header_row(sheet)
    sheet.freeze_panes = 'A2'
    autofit_sheet(sheet)

  raw_sheet = workbook.create_sheet(title='Raw JSON')
  raw_sheet['A1'] = json.dumps(data, indent=2, ensure_ascii=False, default=str)
  raw_sheet.column_dimensions['A'].width = 120
  raw_sheet.row_dimensions[1].height = 400
  raw_sheet['A1'].alignment = Alignment(wrap_text=True, vertical='top')

  summary_sheet.freeze_panes = 'A3'
  summary_sheet.column_dimensions['A'].width = 22
  summary_sheet.column_dimensions['B'].width = 16
  summary_sheet.column_dimensions['D'].width = 24
  summary_sheet.column_dimensions['E'].width = 16
  summary_sheet.column_dimensions['G'].width = 16
  summary_sheet.column_dimensions['H'].width = 16

  buffer = io.BytesIO()
  workbook.save(buffer)
  response = Response(
      buffer.getvalue(),
      media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
  )
  response.headers['Content-Disposition'] = f'attachment; filename=debt-tracker-{today_str}.xlsx'
  return response


@app.route('/import', methods=['POST'])
async def import_data(request: Request):
  form = await read_form_data(request)
  f = form.get('save_file')
  if not f:
    return redirect(url_for('settings_page', status='No file selected.'))
  try:
    raw = (await f.read()).decode('utf-8') if hasattr(f, 'read') else f.file.read().decode('utf-8')
    imported = json.loads(raw)
  except Exception:
    return redirect(url_for('settings_page', status='Invalid file — could not parse JSON.'))
  if not isinstance(imported, dict) or 'accounts' not in imported:
    return redirect(url_for('settings_page', status='Invalid save file — missing accounts data.'))
  save_data(imported)
  count = len(imported.get('accounts', []))
  return redirect(url_for('settings_page', status=f'Imported successfully — {count} account{"s" if count != 1 else ""} loaded.'))


@app.route('/import_excel', methods=['POST'])
async def import_excel(request: Request):
  form = await read_form_data(request)
  f = form.get('excel_file')
  if not f:
    return redirect(url_for('settings_page', status='No Excel file selected.'))

  try:
    raw = await f.read() if hasattr(f, 'read') else f.file.read()
    workbook = load_workbook(io.BytesIO(raw))
  except Exception:
    return redirect(url_for('settings_page', status='Invalid Excel file — could not open workbook.'))

  def sheet_to_dicts(sheet_name):
    if sheet_name not in workbook.sheetnames:
      return []
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
      return []
    headers = [str(h).strip() if h is not None else '' for h in rows[0]]
    result = []
    for row in rows[1:]:
      if not any(cell is not None and str(cell).strip() != '' for cell in row):
        continue
      item = {}
      for idx, header in enumerate(headers):
        if not header:
          continue
        item[header] = coerce_excel_value(row[idx] if idx < len(row) else None)
      result.append(item)
    return result

  accounts_rows = sheet_to_dicts('Accounts')
  history_rows = sheet_to_dicts('History')
  monthly_rows = sheet_to_dicts('Monthly Totals')
  settings_rows = sheet_to_dicts('Settings')

  if not accounts_rows and 'Raw JSON' not in workbook.sheetnames:
    return redirect(url_for('settings_page', status='Invalid workbook — no Accounts sheet found.'))

  data = load_data()
  new_accounts = []

  history_by_index = {}
  history_by_name = {}
  for row in history_rows:
    idx_value = row.get('account_index')
    name_value = row.get('account_name')
    entry = {
      'timestamp': row.get('timestamp') or datetime.utcnow().isoformat(),
      'action': row.get('action') or 'Updated',
      'amount': row.get('amount'),
      'prev_amount': row.get('prev_amount'),
      'payment': row.get('payment'),
      'old_type': row.get('old_type'),
      'new_type': row.get('new_type'),
      'old_name': row.get('old_name'),
      'new_name': row.get('new_name'),
      'old_due_date': row.get('old_due_date'),
      'new_due_date': row.get('new_due_date'),
      'old_interest_rate': row.get('old_interest_rate'),
      'new_interest_rate': row.get('new_interest_rate'),
    }
    if isinstance(idx_value, int):
      history_by_index.setdefault(idx_value, []).append(entry)
    if name_value:
      history_by_name.setdefault(str(name_value), []).append(entry)

  monthly_by_index = {}
  for row in monthly_rows:
    idx_value = row.get('account_index')
    if not isinstance(idx_value, int):
      continue
    monthly_by_index.setdefault(idx_value, []).append({
      'month': row.get('month'),
      'total': row.get('total'),
    })

  for row in accounts_rows:
    account = {}
    for key, value in row.items():
      if key in ('index', 'history_count', 'monthly_value', 'effective_payment'):
        continue
      if key == 'is_archived':
        account['archived'] = bool(value)
        continue
      account[key] = value
    if account.get('archived') is None:
      account.pop('archived', None)
    idx_value = row.get('index')
    history = []
    if isinstance(idx_value, int) and idx_value in history_by_index:
      history = history_by_index[idx_value]
    elif account.get('name') and str(account['name']) in history_by_name:
      history = history_by_name[str(account['name'])]
    account['history'] = sorted(history, key=lambda e: parse_timestamp(e.get('timestamp')) or datetime.min)
    if isinstance(idx_value, int) and idx_value in monthly_by_index:
      account['monthly_totals'] = monthly_by_index[idx_value]
    new_accounts.append(account)

  if not new_accounts and 'Raw JSON' in workbook.sheetnames:
    raw_sheet = workbook['Raw JSON']
    raw_text = raw_sheet['A1'].value
    try:
      imported = json.loads(raw_text) if raw_text else None
      if isinstance(imported, dict) and 'accounts' in imported:
        data = imported
        save_data(data)
        return redirect(url_for('settings_page', status='Imported Excel workbook via Raw JSON sheet.'))
    except Exception:
      pass

  if not new_accounts:
    return redirect(url_for('settings_page', status='Invalid workbook — no account rows could be read.'))

  data['accounts'] = new_accounts
  if settings_rows:
    settings = {}
    for row in settings_rows:
      key = row.get('key')
      if key:
        settings[str(key)] = row.get('value')
    data['settings'] = settings
  save_data(data)
  return redirect(url_for('settings_page', status=f'Imported Excel workbook — {len(new_accounts)} account{"s" if len(new_accounts) != 1 else ""} loaded.'))


@app.route('/download_history', methods=['GET'])
def download_history():
  data = load_data()
  rows = build_history_rows(data)
  fieldnames = [
    'account_name', 'timestamp', 'action', 'amount', 'prev_amount', 'payment',
    'old_type', 'new_type', 'old_name', 'new_name', 'old_due_date', 'new_due_date',
    'old_interest_rate', 'new_interest_rate', 'account_type', 'due_date', 'interest_rate'
  ]
  output = io.StringIO()
  writer = csv.DictWriter(output, fieldnames=fieldnames)
  writer.writeheader()
  for row in rows:
    writer.writerow({k: row.get(k, '') for k in fieldnames})
  response = Response(output.getvalue(), mimetype='text/csv')
  response.headers['Content-Disposition'] = 'attachment; filename=debt_history.csv'
  return response


@app.route('/import_monthly_totals', methods=['POST'])
async def import_monthly_totals(request: Request):
  form = await read_form_data(request)
  raw_text = form.get('monthly_totals', '').strip()
  if not raw_text:
    return redirect(url_for('settings_page', status='No monthly totals entered'))

  lines = [line for line in raw_text.splitlines() if line.strip()]
  if not lines:
    return redirect(url_for('settings_page', status='No monthly totals entered'))

  reader = csv.reader(io.StringIO(raw_text))
  parsed = [row for row in reader if any(cell.strip() for cell in row)]
  if not parsed:
    return redirect(url_for('settings_page', status='No valid monthly totals found'))

  header = [cell.strip().lower() for cell in parsed[0]]
  has_header = len(header) >= 3 and header[0] in ('account_name', 'account', 'name') and header[1] in ('month', 'date', 'timestamp')
  if has_header:
    fieldnames = header
    data_rows = parsed[1:]
  else:
    fieldnames = ['account_name', 'month', 'total', 'type', 'due_date', 'interest_rate']
    data_rows = parsed

  imported = 0
  data = load_data()
  raw_accounts = data.setdefault('accounts', [])

  for row in data_rows:
    if len(row) < 3:
      continue
    row_dict = {fieldnames[i]: row[i].strip() if i < len(row) else '' for i in range(len(fieldnames))}
    account_name = row_dict.get('account_name', '').strip()
    month_value = row_dict.get('month', '').strip()
    total_value = row_dict.get('total', '').strip()
    if not account_name or not month_value or not total_value:
      continue
    timestamp = parse_timestamp(month_value)
    if timestamp is None:
      continue
    try:
      amount = float(total_value)
    except ValueError:
      continue

    new_type = row_dict.get('type', '').strip() or None
    new_due_date = row_dict.get('due_date', '').strip() or None
    new_interest_rate = None
    try:
      interest_value = row_dict.get('interest_rate', '').strip()
      new_interest_rate = float(interest_value) if interest_value else None
    except ValueError:
      new_interest_rate = None

    account = next((a for a in raw_accounts if a.get('name', '') == account_name), None)
    if account is None:
      account = {
        'name': account_name,
        'balance': round(amount, 2),
        'type': new_type or 'Other',
        'due_date': new_due_date or '',
        'interest_rate': new_interest_rate,
        'history': []
      }
      raw_accounts.append(account)
    elif new_type:
      account['type'] = new_type
    elif account.get('type') is None:
      account['type'] = 'Other'
    if new_due_date:
      account['due_date'] = new_due_date
    if new_interest_rate is not None:
      account['interest_rate'] = new_interest_rate

    entry = {
      'timestamp': timestamp.isoformat(),
      'action': 'Imported monthly balance',
      'amount': round(amount, 2),
      'prev_amount': None,
      'payment': None,
      'old_type': None,
      'new_type': account.get('type'),
      'old_name': None,
      'new_name': account.get('name'),
      'old_due_date': None,
      'new_due_date': account.get('due_date', ''),
      'old_interest_rate': None,
      'new_interest_rate': account.get('interest_rate')
    }
    existing_signatures = {make_history_signature(e) for e in account.get('history', [])}
    signature = make_history_signature(entry)
    if signature in existing_signatures:
      continue
    account.setdefault('history', []).append(entry)
    imported += 1

    latest_existing = max(
      (parse_timestamp(e.get('timestamp')) or datetime.min for e in account.get('history', [])),
      default=datetime.min
    )
    if timestamp >= latest_existing:
      account['balance'] = round(amount, 2)

  for account in raw_accounts:
    account['history'] = sorted(
      account.get('history', []),
      key=lambda e: parse_timestamp(e.get('timestamp')) or datetime.min
    )

  data['accounts'] = raw_accounts
  save_data(data)
  return redirect(url_for('settings_page', status=f'Imported {imported} monthly totals'))


@app.route('/history', methods=['GET'])
async def history_page(request: Request):
  data = load_data()
  raw_accounts = data.get('accounts', [])
  history_entries = []
  for account in raw_accounts:
    for entry in account.get('history', []):
      history_entries.append({
        'account_name': entry.get('account_name', account.get('name', 'Unknown')),
        'action': entry.get('action', 'Updated'),
        'amount': entry.get('amount', 0),
        'prev_amount': entry.get('prev_amount'),
        'old_type': entry.get('old_type'),
        'new_type': entry.get('new_type', entry.get('type', account.get('type', 'Other'))),
        'old_name': entry.get('old_name'),
        'new_name': entry.get('new_name'),
        'old_due_date': entry.get('old_due_date'),
        'new_due_date': entry.get('new_due_date'),
        'old_interest_rate': entry.get('old_interest_rate'),
        'new_interest_rate': entry.get('new_interest_rate'),
        'timestamp': entry.get('timestamp', '')
      })
  history_entries = sorted(history_entries, key=lambda e: e['timestamp'], reverse=True)
  return render_template_string(TEMPLATE, page='history', history_entries=history_entries)


UPDATE_TEMPLATE = """
<!doctype html>
<html>
  <head>
    <meta charset="utf-8" />
    <title>Update Balance</title>
    <style>
      body { font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Arial, sans-serif; margin: 2rem; background: #f7f9fc; color: #1b2330; }
      label { display:block; margin-bottom:0.5rem; font-weight:600 }
      input, select { padding:0.5rem; width:200px }
      button { padding:0.6rem 0.9rem; margin-top:0.75rem }
    </style>
  </head>
  <body>
    <h2>Update balance for {{ account.name }}</h2>
    <form method="post">
      <label>Current outstanding: ${{ "%.2f" | format(account.balance) }}</label>
      <label for="new_balance">Current balance</label>
      <input id="new_balance" name="new_balance" type="number" step="0.01" min="0" value="{{ "%.2f" | format(account.balance) }}" required />
      <label for="new_type">Type</label>
      <select id="new_type" name="new_type">
        <option value="Credit Card" {% if account.type == 'Credit Card' %}selected{% endif %}>Credit Card</option>
        <option value="Car Loan" {% if account.type == 'Car Loan' %}selected{% endif %}>Car Loan</option>
        <option value="Mortgage" {% if account.type == 'Mortgage' %}selected{% endif %}>Mortgage</option>
        <option value="Student Loan" {% if account.type == 'Student Loan' %}selected{% endif %}>Student Loan</option>
        <option value="Personal Loan" {% if account.type == 'Personal Loan' %}selected{% endif %}>Personal Loan</option>
        <option value="Other" {% if account.type == 'Other' or not account.type %}selected{% endif %}>Other</option>
      </select>
      <div>
        <button type="submit">Save balance</button>
        <a href="/">Cancel</a>
      </div>
    </form>
  </body>
</html>
"""


@app.route('/update/{idx}', methods=['GET', 'POST'])
async def update_account(idx: int, request: Request):
    data = load_data()
    accounts = data.get('accounts', [])
    if idx < 0 or idx >= len(accounts):
        abort(404)
    account = accounts[idx]
    if request.method == 'POST':
        try:
            json_data = await request.json()
        except Exception:
            json_data = None
        if json_data is not None:
            new_balance = json_data.get('new_balance', account.get('balance', 0))
            new_type = json_data.get('new_type', account.get('type', 'Other'))
            target_month = json_data.get('target_month') or datetime.now().strftime('%Y-%m')
            mark_paid = json_data.get('mark_paid')
            if mark_paid is True:
                account['paid_month'] = target_month
            elif mark_paid is False and account.get('paid_month') == target_month:
                account.pop('paid_month', None)
            mark_paid_2 = json_data.get('mark_paid_2')
            if mark_paid_2 is True:
                account['paid_month_2'] = target_month
            elif mark_paid_2 is False and account.get('paid_month_2') == target_month:
                account.pop('paid_month_2', None)
        else:
            form = await request.form()
            new_balance_value = form.get('new_balance', '')
            try:
                new_balance = float(new_balance_value)
            except (ValueError, TypeError):
                new_balance = account.get('balance', 0)
            new_type = form.get('new_type', account.get('type', 'Other'))

        try:
            new_balance = round(float(new_balance), 2)
        except (ValueError, TypeError):
            new_balance = account.get('balance', 0)

        prev_balance = account.get('balance', 0)
        prev_type = account.get('type', 'Other')
        account['balance'] = max(0.0, new_balance)
        account['type'] = new_type
        account['last_updated'] = datetime.utcnow().isoformat()

        if account.get('category') != 'recurring':
          account['paid_month'] = datetime.now().strftime('%Y-%m')

        action = None
        if new_balance != prev_balance and new_type != prev_type:
          action = 'Updated balance and type'
        elif new_balance != prev_balance:
          action = 'Updated balance'
        elif new_type != prev_type:
          action = 'Updated type'
        if action:
          account.setdefault('history', []).append({
            'action': action,
            'account_name': account.get('name', 'Unknown'),
            'amount': account['balance'],
            'prev_amount': round(prev_balance, 2),
            'payment': None,
            'old_type': prev_type,
            'new_type': new_type,
            'timestamp': datetime.utcnow().isoformat()
          })

        data['accounts'] = accounts
        save_data(data)
        if json_data is not None:
            return jsonify({
              'balance': account['balance'],
              'message': f"Balance updated. Current outstanding is ${account['balance']:.2f}"
            })
            return redirect(url_for('index'))
    return render_template_string(UPDATE_TEMPLATE, account=account)


@app.route('/edit/{idx}', methods=['POST'])
async def edit_account(idx: int, request: Request):
  data = load_data()
  accounts = data.get('accounts', [])
  if idx < 0 or idx >= len(accounts):
    abort(404)
  account = accounts[idx]
  form = await read_form_data(request)
  name = form.get('name', '').strip()
  typ = form.get('type', 'Other')
  old_name = account.get('name', '')
  old_type = account.get('type', 'Other')
  old_due_date = account.get('due_date', '')
  old_interest_rate = account.get('interest_rate')
  old_owner = account.get('owner', '')
  new_owner = form.get('owner', '').strip()
  new_category = form.get('category', account.get('category', 'debt'))
  new_recurring_amount = None
  new_recurring_frequency = None
  new_due_date_2 = None
  if new_category == 'recurring':
    try:
      new_recurring_amount = round(float(form.get('recurring_amount', 0) or 0), 2)
    except ValueError:
      new_recurring_amount = 0.0
    new_recurring_frequency = form.get('recurring_frequency', 'monthly')

  if new_category == 'recurring' and new_recurring_frequency == 'semi-monthly':
    new_due_date = form.get('semi_date_1', '').strip()
    new_due_date_2 = form.get('semi_date_2', '').strip()
  else:
    due_date_type = form.get('due_date_type')
    if due_date_type == 'custom':
      new_due_date = form.get('due_date', '').strip()
    elif due_date_type in ('1st', '15th', 'last', 'unknown'):
      new_due_date = due_date_type
    else:
      new_due_date = form.get('due_date', '').strip()

  interest_rate = None
  try:
    interest_rate = float(form.get('interest_rate', '') or 0)
    if interest_rate == 0:
      interest_rate = None
    else:
      interest_rate = round(interest_rate, 2)
  except ValueError:
    interest_rate = None

  min_payment = None
  if new_category == 'debt':
    try:
      mp = float(form.get('min_payment', '') or 0)
      min_payment = round(mp, 2) if mp > 0 else None
    except ValueError:
      min_payment = None

  new_balance = None
  if new_category == 'debt':
    try:
      nb = float(form.get('balance', '') or 0)
      new_balance = max(0.0, round(nb, 2))
    except ValueError:
      new_balance = None

  old_min_payment = account.get('min_payment')
  old_balance = account.get('balance', 0)

  changed = False
  action = None
  if name and name != old_name:
    changed = True
  if typ != old_type:
    changed = True
  if new_due_date != old_due_date:
    changed = True
  if interest_rate != old_interest_rate:
    changed = True
  if new_owner != old_owner:
    changed = True
  if new_category != account.get('category', 'debt'):
    changed = True
  if min_payment != old_min_payment:
    changed = True
  if new_balance is not None and new_balance != old_balance:
    changed = True
  if changed:
    if name != old_name and typ != old_type:
      action = 'Updated name and type'
    elif name != old_name:
      action = 'Updated name'
    elif typ != old_type:
      action = 'Updated type'
    elif new_due_date != old_due_date and interest_rate != old_interest_rate:
      action = 'Updated due date and interest rate'
    elif new_due_date != old_due_date:
      action = 'Updated due date'
    elif interest_rate != old_interest_rate:
      action = 'Updated interest rate'
    elif new_owner != old_owner:
      action = 'Updated owner'
    else:
      action = 'Updated account'
    account['name'] = name or old_name
    account['type'] = typ
    account['owner'] = new_owner
    account['due_date'] = new_due_date
    account['category'] = new_category
    account['due_date_2'] = new_due_date_2 if new_due_date_2 else None
    if new_category == 'recurring':
      account['recurring_amount'] = new_recurring_amount
      account['recurring_frequency'] = new_recurring_frequency
      account['interest_rate'] = None
      account.pop('min_payment', None)
    else:
      account['interest_rate'] = interest_rate
      account['min_payment'] = min_payment
      if new_balance is not None:
        account['balance'] = new_balance
      account.pop('recurring_amount', None)
      account.pop('recurring_frequency', None)
      account.pop('due_date_2', None)
    account.setdefault('history', []).append({
      'action': action,
      'account_name': account.get('name', 'Unknown'),
      'amount': account.get('balance', 0),
      'prev_amount': None,
      'old_name': old_name,
      'new_name': account.get('name', old_name),
      'old_type': old_type,
      'new_type': typ,
      'old_due_date': old_due_date,
      'new_due_date': new_due_date,
      'old_interest_rate': old_interest_rate,
      'new_interest_rate': interest_rate,
      'timestamp': datetime.utcnow().isoformat()
    })
  data['accounts'] = accounts
  save_data(data)
  if form.get('return_to') == 'dashboard':
    return redirect(url_for('index'))
  return redirect(url_for('settings_page'))


@app.route('/archive/{idx}', methods=['POST'])
async def archive_account(idx: int, request: Request):
  data = load_data()
  accounts = data.get('accounts', [])
  if 0 <= idx < len(accounts):
    accounts[idx]['archived'] = True
    save_data(data)
  return ('', 200)


@app.route('/unarchive/{idx}', methods=['POST'])
async def unarchive_account(idx: int, request: Request):
  data = load_data()
  accounts = data.get('accounts', [])
  if 0 <= idx < len(accounts):
    accounts[idx]['archived'] = False
    save_data(data)
  return ('', 200)


@app.route('/delete/{idx}', methods=['POST'])
async def delete_account(idx: int, request: Request):
  data = load_data()
  accounts = data.get('accounts', [])
  if idx < 0 or idx >= len(accounts):
    abort(404)
  # remove the account
  accounts.pop(idx)
  data['accounts'] = accounts
  save_data(data)
  return ('', 200)


def find_free_port(start_port=5000, max_port=5010):
    for port in range(start_port, max_port + 1):
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            s.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
            try:
                s.bind(("0.0.0.0", port))
                return port
            except OSError:
                continue
    raise RuntimeError("No available ports found")


def open_browser(url):
    webbrowser.open(url)

if __name__ == "__main__":
    init_storage()
    port = int(os.environ.get("PORT", find_free_port(5000, 5010)))
    url = f"http://127.0.0.1:{port}"
    threading.Timer(1.0, lambda: open_browser(url)).start()
    print(f"Starting Debt Tracker on {url}")
    uvicorn.run(app, host="0.0.0.0", port=port, reload=False, log_level="debug")
